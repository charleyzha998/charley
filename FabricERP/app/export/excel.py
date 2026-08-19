"""Excel 导出（openpyxl）：送货单、对账单、库存表。"""

import os
from tkinter import filedialog

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .. import models
from ..db import get_setting

TITLE_FONT = Font(name="微软雅黑", size=16, bold=True)
HEAD_FONT = Font(name="微软雅黑", size=10, bold=True)
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD = Font(name="微软雅黑", size=10, bold=True)
HEAD_FILL = PatternFill("solid", fgColor="DCE6F1")
THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _ask_path(default_name, parent=None):
    return filedialog.asksaveasfilename(
        parent=parent, title="导出到", defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")])


def _write_title(ws, text, ncols, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text)
    c.font = TITLE_FONT
    c.alignment = CENTER
    ws.row_dimensions[1].height = 30
    row = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws.cell(2, 1, subtitle)
        c.font = BODY_FONT
        c.alignment = LEFT
        row = 3
    return row


def _write_header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row, i, h)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    return row + 1


def _write_row(ws, row, values, aligns, bold=False):
    for i, (v, a) in enumerate(zip(values, aligns), start=1):
        c = ws.cell(row, i, v)
        c.font = BOLD if bold else BODY_FONT
        c.alignment = {"c": CENTER, "r": RIGHT, "l": LEFT}[a]
        c.border = BORDER
        if isinstance(v, float):
            c.number_format = "#,##0.00"
    return row + 1


def _company_line():
    parts = [get_setting("company_name"), get_setting("company_phone")]
    return "　".join(p for p in parts if p)


# ---------------- 送货单 ----------------

def export_delivery(shipment_id, parent=None):
    head, items = models.get_shipment(shipment_id)
    path = _ask_path(f"送货单_{head['customer']}_{head['doc_no']}.xlsx", parent)
    if not path:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "送货单"
    headers = ["缸号", "面料名称", "颜色", "工艺", "卷数", "米数", "单价", "金额", "备注"]
    widths = [12, 22, 10, 12, 8, 11, 9, 13, 16]
    aligns = ["c", "l", "c", "c", "r", "r", "r", "r", "l"]

    title = (get_setting("company_name") or "") + " 送货单"
    row = _write_title(ws, title.strip(), len(headers),
                       f"客户：{head['customer']}　　单号：{head['doc_no']}　　"
                       f"日期：{head['ship_date']}　　收货人：{head['receiver'] or ''}　　"
                       f"车牌：{head['plate_no'] or ''}")
    row = _write_header(ws, row, headers, widths)
    for it in items:
        row = _write_row(ws, row, [
            it["dye_lot"], it["fabric"], it["color"], it["process"], it["rolls"],
            float(it["meters"]), float(it["unit_price"]), float(it["amount"]),
            it["note"] or ""], aligns)

    row = _write_row(ws, row, [
        "合计", "", "", "", sum(i["rolls"] for i in items),
        float(sum(i["meters"] for i in items)), "",
        float(sum(i["amount"] for i in items)), ""], aligns, bold=True)

    row += 1
    ws.cell(row, 1, f"备注：{head['note'] or ''}").font = BODY_FONT
    row += 2
    ws.cell(row, 1, "送货人签字：").font = BODY_FONT
    ws.cell(row, 5, "收货人签字：").font = BODY_FONT
    row += 1
    ws.cell(row, 1, _company_line()).font = BODY_FONT

    wb.save(path)
    return path


# ---------------- 对账单 ----------------

def export_statement(st, parent=None):
    cust = st["customer"]
    period = f"{st['date_from'] or '开始'} 至 {st['date_to'] or '至今'}"
    path = _ask_path(f"对账单_{cust['name']}_{(st['date_to'] or '')}.xlsx", parent)
    if not path:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "对账单"
    headers = ["日期", "送货单号", "缸号", "面料名称", "颜色", "工艺",
               "卷数", "米数", "单价", "金额"]
    widths = [12, 16, 12, 22, 10, 12, 8, 11, 9, 13]
    aligns = ["c", "c", "c", "l", "c", "c", "r", "r", "r", "r"]

    title = (get_setting("company_name") or "") + " 对账单"
    row = _write_title(ws, title.strip(), len(headers),
                       f"客户：{cust['name']}　　对账区间：{period}")
    row = _write_header(ws, row, headers, widths)

    for it in st["items"]:
        row = _write_row(ws, row, [
            it["ship_date"], it["doc_no"], it["dye_lot"], it["fabric"], it["color"],
            it["process"], it["rolls"], float(it["meters"]),
            float(it["unit_price"]), float(it["amount"])], aligns)

    row = _write_row(ws, row, [
        "合计", "", "", f"{len(st['items'])} 行", "", "", st["total_rolls"],
        float(st["total_meters"]), "", float(st["billed"])], aligns, bold=True)

    # 收款明细
    if st["payments"]:
        row += 1
        ws.cell(row, 1, "本期收款").font = HEAD_FONT
        row += 1
        row = _write_header(ws, row, ["日期", "金额", "方式", "凭证号", "备注"],
                            [12, 13, 10, 18, 24])
        for p in st["payments"]:
            row = _write_row(ws, row, [p["pay_date"], float(p["amount"]), p["method"],
                                       p["ref_no"] or "", p["note"] or ""],
                             ["c", "r", "c", "l", "l"])

    # 结算汇总
    row += 1
    for label, val in [("期初欠款", st["opening"]), ("本期应收（加工费）", st["billed"]),
                       ("本期已收", st["paid"]), ("期末应收（客户欠款）", st["closing"])]:
        ws.cell(row, 8, label).font = BOLD
        ws.cell(row, 8).alignment = RIGHT
        c = ws.cell(row, 10, float(val))
        c.font = BOLD
        c.number_format = "#,##0.00"
        c.alignment = RIGHT
        c.border = BORDER
        row += 1

    row += 1
    bank = get_setting("company_bank")
    if bank:
        ws.cell(row, 1, f"收款账号：{bank}").font = BODY_FONT
        row += 1
    ws.cell(row, 1, _company_line()).font = BODY_FONT

    wb.save(path)
    return path


# ---------------- 库存表 ----------------

def export_stock(rows, parent=None, customer=None):
    """缸号库存表：坯布 / 成品 / 已发 三段，发给客户对库存用。"""
    name = f"库存表_{customer}.xlsx" if customer else "库存表.xlsx"
    path = _ask_path(name, parent)
    if not path:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "库存"
    headers = ["客户", "进仓日期", "缸号", "面料名称", "颜色", "进仓卷", "进仓米",
               "未加工卷", "未加工米", "待发卷", "待发米", "已发卷", "已发米",
               "状态", "缩率%", "备注"]
    widths = [16, 12, 12, 20, 10, 8, 11, 9, 11, 8, 11, 8, 11, 10, 9, 16]
    aligns = ["l", "c", "c", "l", "c", "r", "r", "r", "r", "r", "r", "r", "r",
              "c", "r", "l"]

    title = (get_setting("company_name") or "") + " 面料库存表"
    sub = f"客户：{customer}" if customer else None
    row = _write_title(ws, title.strip(), len(headers), sub)
    row = _write_header(ws, row, headers, widths)
    for r in rows:
        row = _write_row(ws, row, [
            r["customer"], r["in_date"], r["dye_lot"], r["fabric"], r["color"],
            r["in_rolls"], float(r["in_meters"]),
            r["greige_rolls"], float(r["greige_meters"]),
            r["fin_rolls"], float(r["fin_meters"]),
            r["out_rolls"], float(r["out_meters"]), r["state"],
            "" if r["shrink_pct"] is None else float(r["shrink_pct"]),
            r["note"] or ""], aligns)
    row = _write_row(ws, row, [
        "合计", "", "", f"{len(rows)} 缸", "",
        sum(r["in_rolls"] for r in rows), float(sum(r["in_meters"] for r in rows)),
        sum(r["greige_rolls"] for r in rows),
        float(sum(r["greige_meters"] for r in rows)),
        sum(r["fin_rolls"] for r in rows), float(sum(r["fin_meters"] for r in rows)),
        sum(r["out_rolls"] for r in rows), float(sum(r["out_meters"] for r in rows)),
        "", "", ""], aligns, bold=True)

    row += 1
    ws.cell(row, 1, "说明：未加工 = 进仓了还没做的坯布；待发 = 做好了还压在厂里的成品；"
                    "已发完以卷数为准（染整有缩率，米数不会正好归零）。").font = BODY_FONT
    row += 1
    ws.cell(row, 1, _company_line()).font = BODY_FONT
    wb.save(path)
    return path


def export_finished(rows, parent=None, customer=None):
    """成品库存表：加工好了还没发的货。"""
    name = f"成品库存_{customer}.xlsx" if customer else "成品库存表.xlsx"
    path = _ask_path(name, parent)
    if not path:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "成品库存"
    headers = ["客户", "加工日期", "缸号", "面料名称", "颜色", "工艺",
               "成品卷", "成品米", "已发卷", "已发米", "待发卷", "待发米",
               "状态", "备注"]
    widths = [16, 12, 12, 20, 10, 12, 8, 11, 8, 11, 8, 11, 10, 16]
    aligns = ["l", "c", "c", "l", "c", "c", "r", "r", "r", "r", "r", "r", "c", "l"]

    title = (get_setting("company_name") or "") + " 成品库存表（已加工待发）"
    sub = f"客户：{customer}" if customer else None
    row = _write_title(ws, title.strip(), len(headers), sub)
    row = _write_header(ws, row, headers, widths)
    for r in rows:
        row = _write_row(ws, row, [
            r["customer"], r["done_date"], r["dye_lot"], r["fabric"], r["color"],
            r["process"], r["done_rolls"], float(r["done_meters"]),
            r["out_rolls"], float(r["out_meters"]),
            r["left_rolls"], float(r["left_meters"]), r["state"],
            r["note"] or ""], aligns)
    row = _write_row(ws, row, [
        "合计", "", "", f"{len(rows)} 条", "", "",
        sum(r["done_rolls"] for r in rows), float(sum(r["done_meters"] for r in rows)),
        sum(r["out_rolls"] for r in rows), float(sum(r["out_meters"] for r in rows)),
        sum(r["left_rolls"] for r in rows), float(sum(r["left_meters"] for r in rows)),
        "", ""], aligns, bold=True)
    row += 2
    ws.cell(row, 1, _company_line()).font = BODY_FONT
    wb.save(path)
    return path

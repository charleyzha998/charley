"""Excel 批量导入：客户 / 价格表 / 进仓 / 发货 / 收款。

给会计用的：生成模板 → 会计在 Excel 里填 → 拿回来一键导入。
会计电脑上不用装本软件，也就不存在两份数据打架的问题。

设计原则：
- 先全量校验（analyze），一条错都没有才允许导入 —— 不做「导一半失败」
- 整个导入在一个事务里，中途出错全部回滚
- 导入前自动备份
"""

import datetime
import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import backup
from .db import get_conn, transaction
from .services import next_doc_no

# ---------------- 模板定义 ----------------
# (列名, 是否必填, 列宽)
SHEETS = {
    "客户": [("客户名称", 1, 20), ("客户编号", 0, 12), ("联系人", 0, 12),
             ("电话", 0, 16), ("地址", 0, 28), ("期初欠款", 0, 12),
             ("期初日期", 0, 12), ("备注", 0, 20)],
    "价格表": [("客户名称", 1, 20), ("面料", 0, 18), ("工艺", 1, 12),
               ("单价", 1, 10), ("生效日期", 0, 12)],
    "进仓": [("客户名称", 1, 20), ("进仓日期", 1, 12), ("缸号", 1, 14),
             ("面料", 0, 18), ("颜色", 0, 12), ("卷数", 1, 8),
             ("米数", 1, 10), ("备注", 0, 20)],
    "发货": [("客户名称", 1, 20), ("发货日期", 1, 12), ("缸号", 1, 14),
             ("工艺", 0, 12), ("卷数", 1, 8), ("米数", 1, 10),
             ("单价", 0, 10), ("备注", 0, 20)],
    "收款": [("客户名称", 1, 20), ("收款日期", 1, 12), ("金额", 1, 12),
             ("方式", 0, 10), ("单据号", 0, 16), ("备注", 0, 20)],
}

SHEET_ORDER = ["客户", "价格表", "进仓", "发货", "收款"]

SAMPLES = {
    "客户": [["宁波华丰纺织", "HF01", "王经理", "13800138000",
              "浙江省宁波市…", 5000, "2026-01-01", "老客户"]],
    "价格表": [["宁波华丰纺织", "涤纶四面弹", "贴白膜", 3.5, "2026-01-01"],
               ["宁波华丰纺织", "", "贴黑膜", 0.8, "2026-01-01"]],
    "进仓": [["宁波华丰纺织", "2026-07-05", "D2601", "涤纶四面弹", "藏青",
              20, 1000, ""],
             ["宁波华丰纺织", "2026-07-05", "D2602", "涤纶四面弹", "米白",
              15, 780, ""]],
    "发货": [["宁波华丰纺织", "2026-07-12", "D2601", "贴白膜", 12, 585, 3.5, ""]],
    "收款": [["宁波华丰纺织", "2026-07-20", 20000, "转账", "", ""]],
}

NOTES = {
    "客户": "期初欠款 = 用本系统之前客户还欠的加工费；不填按 0 算。客户名称不能重复。",
    "价格表": "面料留空 = 这个客户这道工艺的通用价。发货时按「客户+面料+工艺」自动带出单价。",
    "进仓": "缸号在同一个客户里不能重复。同一客户同一天的多行会自动合成一张进仓单。",
    "发货": "缸号必须在「进仓」表里有，或者系统里已经存在。单价留空会自动去价格表里找。",
    "收款": "方式可填：现金 / 转账 / 承兑 / 抵扣 / 其他，不填按「转账」算。",
}

HELP_TEXT = [
    ("面料复合加工管理系统 —— 数据导入模板", True),
    ("", False),
    ("怎么用：", True),
    ("1. 下面每个工作表（客户、价格表、进仓、发货、收款）按需要填，", False),
    ("   不需要的整张表留空就行，不用勉强填。", False),
    ("2. 红色带 * 的列是必填，其余可以空着。", False),
    ("3. 「填写示例」这张表是给你看格式的，不会被导入，不用管它。", False),
    ("4. 填好保存，把文件交回，在软件里点「导入数据」选这个文件。", False),
    ("", False),
    ("填写规则：", True),
    ("· 日期写成 2026-07-05 这种格式；Excel 的日期格式也认。", False),
    ("· 卷数填整数，米数和金额填数字，不要带「米」「元」这些字。", False),
    ("· 客户名称必须前后一致，多一个空格或写法不同会被当成两个客户。", False),
    ("· 面料和工艺如果系统里没有，导入时会自动建档，不用提前录。", False),
    ("", False),
    ("建议（很重要）：", True),
    ("· 历史的发货和收款不用全部补录。", False),
    ("  最省事的做法是：把每个客户到今天为止还欠多少，填在「客户」表的", False),
    ("  「期初欠款」里，然后只导入目前还没发完的缸号（进仓表）。", False),
    ("  这样账目从今天开始就是准的，也不用翻几年的老账。", False),
    ("· 「发货」「收款」两张表是给想补录历史明细的人用的，可以整张留空。", False),
    ("", False),
    ("导入前软件会先自动备份，并且会把数据从头到尾检查一遍，", False),
    ("有问题会列出来告诉你在第几行，全部没问题才会真正写入。", False),
]

HDR_FILL = PatternFill("solid", fgColor="DCE6F1")
REQ_FONT = Font(bold=True, color="C00000", name="微软雅黑", size=10)
OPT_FONT = Font(bold=True, name="微软雅黑", size=10)
BODY_FONT = Font(name="微软雅黑", size=10)
GRAY_FONT = Font(name="微软雅黑", size=10, color="808080", italic=True)
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_template(path):
    """生成空白模板（含说明和示例）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    ws.column_dimensions["A"].width = 78
    for i, (text, bold) in enumerate(HELP_TEXT, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name="微软雅黑", size=11, bold=bold)

    for name in SHEET_ORDER:
        sh = wb.create_sheet(name)
        cols = SHEETS[name]
        sh.cell(row=1, column=1, value=NOTES[name]).font = Font(
            name="微软雅黑", size=10, color="808080")
        for j, (title, req, width) in enumerate(cols, start=1):
            c = sh.cell(row=2, column=j, value=title + (" *" if req else ""))
            c.font = REQ_FONT if req else OPT_FONT
            c.fill = HDR_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center")
            sh.column_dimensions[get_column_letter(j)].width = width
        sh.freeze_panes = "A3"

    sh = wb.create_sheet("填写示例")
    sh.cell(row=1, column=1, value="这张表只是给你看格式用的，不会被导入。").font = Font(
        name="微软雅黑", size=11, bold=True, color="C00000")
    r = 3
    for name in SHEET_ORDER:
        sh.cell(row=r, column=1, value=f"【{name}】").font = OPT_FONT
        r += 1
        for j, (title, req, width) in enumerate(SHEETS[name], start=1):
            c = sh.cell(row=r, column=j, value=title + (" *" if req else ""))
            c.font = REQ_FONT if req else OPT_FONT
            c.fill = HDR_FILL
            c.border = BORDER
        r += 1
        for sample in SAMPLES[name]:
            for j, v in enumerate(sample, start=1):
                c = sh.cell(row=r, column=j, value=v)
                c.font = GRAY_FONT
                c.border = BORDER
            r += 1
        r += 1
    for j in range(1, 9):
        sh.column_dimensions[get_column_letter(j)].width = 18

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    return path


# ---------------- 取值 ----------------

DATE_RE = re.compile(r"^(\d{4})[-/.]?(\d{1,2})[-/.]?(\d{1,2})$")


def _text(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _date(v):
    """接受 Excel 日期格式、2026-07-05、2026/7/5、20260705。"""
    if v is None or _text(v) == "":
        return None
    if isinstance(v, datetime.datetime):
        return v.date().strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    m = DATE_RE.match(_text(v))
    if not m:
        raise ValueError(f"日期「{_text(v)}」看不懂，应该写成 2026-07-05")
    y, mo, d = (int(g) for g in m.groups())
    try:
        return datetime.date(y, mo, d).strftime("%Y-%m-%d")
    except ValueError:
        # 2026-13-45 这种：格式对得上，但月/日超范围。别把 Python 的英文报错抛给会计
        raise ValueError(f"日期「{_text(v)}」不存在，请检查月份和日号")


def _num(v, default=None):
    s = _text(v).replace(",", "").replace("，", "")
    for unit in ("米", "元", "卷", "㎡"):
        s = s.replace(unit, "")
    s = s.strip()
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        raise ValueError(f"「{_text(v)}」不是数字")


def _int(v, default=None):
    n = _num(v, None)
    if n is None:
        return default
    if abs(n - round(n)) > 1e-6:
        raise ValueError(f"「{_text(v)}」必须是整数")
    return int(round(n))


def read_book(path):
    """读出每张表的数据行。返回 {表名: [ {列名: 值, '_row': 行号}, ... ]}。"""
    wb = load_workbook(path, data_only=True)
    out = {}
    for name in SHEET_ORDER:
        rows = []
        if name not in wb.sheetnames:
            out[name] = rows
            continue
        sh = wb[name]
        titles = [t for t, _, _ in SHEETS[name]]
        # 找表头行：哪一行含有第一个必填列名
        head_row, colmap = None, {}
        for r in range(1, min(sh.max_row, 10) + 1):
            vals = {}
            for c in range(1, sh.max_column + 1):
                t = _text(sh.cell(row=r, column=c).value).rstrip("*").strip()
                if t in titles:
                    vals[t] = c
            if titles[0] in vals:
                head_row, colmap = r, vals
                break
        if head_row is None:
            out[name] = rows
            continue
        for r in range(head_row + 1, sh.max_row + 1):
            row = {t: sh.cell(row=r, column=c).value for t, c in colmap.items()}
            if all(_text(v) == "" for v in row.values()):
                continue
            row["_row"] = r
            rows.append(row)
        out[name] = rows
    wb.close()
    return out


# ---------------- 校验 ----------------

class Report:
    def __init__(self):
        self.errors = []     # [(表, 行, 说明)] —— 有这个就不能导
        self.warnings = []   # [(表, 行, 说明)] —— 可以导，但提醒看一眼
        self.stats = {}
        self.plan = None

    def err(self, sheet, row, msg):
        self.errors.append((sheet, row, msg))

    def warn(self, sheet, row, msg):
        self.warnings.append((sheet, row, msg))

    @property
    def ok(self):
        return not self.errors and any(self.stats.values())


def analyze(path):
    """全量校验，不写任何数据。返回 Report。"""
    rep = Report()
    try:
        book = read_book(path)
    except Exception as e:
        rep.err("文件", 0, f"打不开这个 Excel：{e}")
        return rep

    conn = get_conn()
    db_customers = {r["name"]: r["id"] for r in
                    conn.execute("SELECT id, name FROM customer")}
    db_lots = {(r["customer_id"], r["dye_lot"]): r["id"] for r in
               conn.execute("SELECT id, customer_id, dye_lot FROM inbound_item")}
    db_fabrics = {r["name"] for r in conn.execute("SELECT name FROM fabric")}
    db_processes = {r["name"] for r in conn.execute("SELECT name FROM process")}

    plan = {"customers": [], "prices": [], "inbounds": {}, "shipments": {},
            "payments": []}
    new_names, new_fabrics, new_processes = [], set(), set()

    # ---- 客户 ----
    seen = {}
    for row in book["客户"]:
        n, name = row["_row"], _text(row.get("客户名称"))
        if not name:
            rep.err("客户", n, "客户名称是空的")
            continue
        if name in seen:
            rep.err("客户", n, f"「{name}」在第 {seen[name]} 行已经有了，本表里不能重复")
            continue
        seen[name] = n
        if name in db_customers:
            rep.warn("客户", n, f"「{name}」系统里已经有了，这行会跳过，不会覆盖原有资料")
            continue
        try:
            ob = _num(row.get("期初欠款"), 0.0)
            od = _date(row.get("期初日期"))
        except ValueError as e:
            rep.err("客户", n, str(e))
            continue
        plan["customers"].append({
            "name": name, "code": _text(row.get("客户编号")) or None,
            "contact": _text(row.get("联系人")) or None,
            "phone": _text(row.get("电话")) or None,
            "address": _text(row.get("地址")) or None,
            "opening_balance": ob, "opening_date": od,
            "note": _text(row.get("备注")) or None})
        new_names.append(name)

    known = set(db_customers) | set(new_names)

    def need_customer(sheet, n, name):
        if not name:
            rep.err(sheet, n, "客户名称是空的")
            return False
        if name not in known:
            rep.err(sheet, n, f"客户「{name}」既不在「客户」表里，系统里也没有。"
                              f"请检查名字是不是写得不一样")
            return False
        return True

    # ---- 价格表 ----
    pkeys = {}
    for row in book["价格表"]:
        n, name = row["_row"], _text(row.get("客户名称"))
        if not need_customer("价格表", n, name):
            continue
        proc = _text(row.get("工艺"))
        if not proc:
            rep.err("价格表", n, "工艺不能空")
            continue
        try:
            price = _num(row.get("单价"))
            eff = _date(row.get("生效日期")) or datetime.date.today().strftime("%Y-%m-%d")
        except ValueError as e:
            rep.err("价格表", n, str(e))
            continue
        if price is None:
            rep.err("价格表", n, "单价不能空")
            continue
        if price <= 0:
            rep.warn("价格表", n, f"单价是 {price}，确认一下是不是填错了")
        fab = _text(row.get("面料"))
        key = (name, fab, proc, eff)
        if key in pkeys:
            rep.err("价格表", n, f"和第 {pkeys[key]} 行重复了"
                                 f"（同客户+同面料+同工艺+同生效日期只能有一条）")
            continue
        pkeys[key] = n
        if fab and fab not in db_fabrics:
            new_fabrics.add(fab)
        if proc not in db_processes:
            new_processes.add(proc)
        plan["prices"].append({"customer": name, "fabric": fab or None,
                               "process": proc, "price": price, "eff": eff})

    # ---- 进仓 ----
    sheet_lots = {}
    for row in book["进仓"]:
        n, name = row["_row"], _text(row.get("客户名称"))
        if not need_customer("进仓", n, name):
            continue
        lot = _text(row.get("缸号"))
        if not lot:
            rep.err("进仓", n, "缸号不能空")
            continue
        if (name, lot) in sheet_lots:
            rep.err("进仓", n, f"缸号「{lot}」和第 {sheet_lots[(name, lot)]} 行重复了。"
                               f"同一个客户的缸号必须唯一")
            continue
        if name in db_customers and (db_customers[name], lot) in db_lots:
            rep.err("进仓", n, f"缸号「{lot}」系统里已经有了，不能重复导入")
            continue
        try:
            d = _date(row.get("进仓日期"))
            rolls = _int(row.get("卷数"))
            meters = _num(row.get("米数"))
        except ValueError as e:
            rep.err("进仓", n, str(e))
            continue
        if not d:
            rep.err("进仓", n, "进仓日期不能空")
            continue
        if rolls is None or meters is None:
            rep.err("进仓", n, "卷数和米数都要填")
            continue
        if rolls <= 0 or meters <= 0:
            rep.err("进仓", n, f"卷数({rolls})和米数({meters})要大于 0")
            continue
        sheet_lots[(name, lot)] = n
        fab = _text(row.get("面料"))
        if fab and fab not in db_fabrics:
            new_fabrics.add(fab)
        plan["inbounds"].setdefault((name, d), []).append({
            "dye_lot": lot, "fabric": fab or None,
            "color": _text(row.get("颜色")) or None, "rolls": rolls,
            "meters": round(meters, 2), "note": _text(row.get("备注")) or None,
            "_row": n})

    # ---- 发货 ----
    shipped = {}
    for row in book["发货"]:
        n, name = row["_row"], _text(row.get("客户名称"))
        if not need_customer("发货", n, name):
            continue
        lot = _text(row.get("缸号"))
        in_sheet = (name, lot) in sheet_lots
        in_db = name in db_customers and (db_customers[name], lot) in db_lots
        if not lot:
            rep.err("发货", n, "缸号不能空")
            continue
        if not (in_sheet or in_db):
            rep.err("发货", n, f"缸号「{lot}」在「进仓」表里找不到，系统里也没有。"
                               f"发货的缸号必须先有进仓记录")
            continue
        try:
            d = _date(row.get("发货日期"))
            rolls = _int(row.get("卷数"))
            meters = _num(row.get("米数"))
            price = _num(row.get("单价"), None)
        except ValueError as e:
            rep.err("发货", n, str(e))
            continue
        if not d:
            rep.err("发货", n, "发货日期不能空")
            continue
        if rolls is None or meters is None:
            rep.err("发货", n, "卷数和米数都要填")
            continue
        proc = _text(row.get("工艺"))
        if proc and proc not in db_processes:
            new_processes.add(proc)
        if price is None and not proc:
            rep.warn("发货", n, "没填单价也没填工艺，这行金额会算成 0")
        shipped[(name, lot)] = shipped.get((name, lot), 0) + rolls
        if in_sheet:
            avail = next(it["rolls"] for its in plan["inbounds"].values()
                         for it in its if it["dye_lot"] == lot)
            if shipped[(name, lot)] > avail:
                rep.warn("发货", n, f"缸号「{lot}」累计发了 {shipped[(name, lot)]} 卷，"
                                    f"进仓只有 {avail} 卷，超发了。能导，但请核对")
        plan["shipments"].setdefault((name, d), []).append({
            "dye_lot": lot, "process": proc or None, "rolls": rolls,
            "meters": round(meters, 2), "price": price,
            "note": _text(row.get("备注")) or None, "_row": n})

    # ---- 收款 ----
    for row in book["收款"]:
        n, name = row["_row"], _text(row.get("客户名称"))
        if not need_customer("收款", n, name):
            continue
        try:
            d = _date(row.get("收款日期"))
            amt = _num(row.get("金额"))
        except ValueError as e:
            rep.err("收款", n, str(e))
            continue
        if not d:
            rep.err("收款", n, "收款日期不能空")
            continue
        if amt is None:
            rep.err("收款", n, "金额不能空")
            continue
        if amt <= 0:
            rep.warn("收款", n, f"金额是 {amt}，确认一下是不是填错了")
        plan["payments"].append({
            "customer": name, "date": d, "amount": round(amt, 2),
            "method": _text(row.get("方式")) or "转账",
            "ref_no": _text(row.get("单据号")) or None,
            "note": _text(row.get("备注")) or None})

    rep.plan = plan
    rep.stats = {
        "客户": len(plan["customers"]),
        "价格": len(plan["prices"]),
        "进仓单": len(plan["inbounds"]),
        "进仓缸号": sum(len(v) for v in plan["inbounds"].values()),
        "发货单": len(plan["shipments"]),
        "发货明细": sum(len(v) for v in plan["shipments"].values()),
        "收款": len(plan["payments"]),
    }
    rep.new_fabrics = sorted(new_fabrics)
    rep.new_processes = sorted(new_processes)
    if not any(rep.stats.values()) and not rep.errors:
        rep.err("文件", 0, "这个文件里没有可导入的数据，检查一下是不是填错了工作表")
    return rep


# ---------------- 执行 ----------------

def run_import(rep):
    """按校验过的 plan 写库。整个过程一个事务，出错全部回滚。"""
    if not rep.ok:
        raise ValueError("还有错误没改完，不能导入。")
    backup.backup_now("before_import")
    plan = rep.plan
    done = {k: 0 for k in ("客户", "价格", "进仓单", "缸号", "发货单", "发货明细", "收款")}

    with transaction() as conn:
        cust = {r["name"]: r["id"] for r in conn.execute("SELECT id, name FROM customer")}
        fabs = {r["name"]: r["id"] for r in conn.execute("SELECT id, name FROM fabric")}
        procs = {r["name"]: r["id"] for r in conn.execute("SELECT id, name FROM process")}
        lots = {(r["customer_id"], r["dye_lot"]): r["id"] for r in
                conn.execute("SELECT id, customer_id, dye_lot FROM inbound_item")}

        def fabric_id(name):
            if not name:
                return None
            if name not in fabs:
                fabs[name] = conn.execute(
                    "INSERT INTO fabric(name) VALUES (?)", (name,)).lastrowid
            return fabs[name]

        def process_id(name):
            if not name:
                return None
            if name not in procs:
                procs[name] = conn.execute(
                    "INSERT INTO process(name) VALUES (?)", (name,)).lastrowid
            return procs[name]

        for c in plan["customers"]:
            cust[c["name"]] = conn.execute(
                """INSERT INTO customer(name, code, contact, phone, address,
                   opening_balance, opening_date, note) VALUES (?,?,?,?,?,?,?,?)""",
                (c["name"], c["code"], c["contact"], c["phone"], c["address"],
                 c["opening_balance"], c["opening_date"], c["note"])).lastrowid
            done["客户"] += 1

        for p in plan["prices"]:
            conn.execute(
                """INSERT INTO price(customer_id, fabric_id, process_id,
                   unit_price, effective_date) VALUES (?,?,?,?,?)""",
                (cust[p["customer"]], fabric_id(p["fabric"]),
                 process_id(p["process"]), p["price"], p["eff"]))
            done["价格"] += 1

        for (name, d), items in sorted(plan["inbounds"].items()):
            cid = cust[name]
            iid = conn.execute(
                "INSERT INTO inbound(doc_no, customer_id, in_date, note) VALUES (?,?,?,?)",
                (next_doc_no("JC", d, conn), cid, d, "Excel 导入")).lastrowid
            done["进仓单"] += 1
            for it in items:
                lots[(cid, it["dye_lot"])] = conn.execute(
                    """INSERT INTO inbound_item(inbound_id, customer_id, dye_lot,
                       fabric_id, color, rolls, meters, note)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (iid, cid, it["dye_lot"], fabric_id(it["fabric"]), it["color"],
                     it["rolls"], it["meters"], it["note"])).lastrowid
                done["缸号"] += 1

        for (name, d), items in sorted(plan["shipments"].items()):
            cid = cust[name]
            sid = conn.execute(
                """INSERT INTO shipment(doc_no, customer_id, ship_date, note)
                   VALUES (?,?,?,?)""",
                (next_doc_no("FH", d, conn), cid, d, "Excel 导入")).lastrowid
            done["发货单"] += 1
            for it in items:
                item_id = lots[(cid, it["dye_lot"])]
                pid = process_id(it["process"])
                price = it["price"]
                if price is None:
                    price = _lookup_price(conn, cid, item_id, pid, d) or 0.0
                conn.execute(
                    """INSERT INTO shipment_item(shipment_id, inbound_item_id,
                       process_id, rolls, meters, unit_price, amount, note)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (sid, item_id, pid, it["rolls"], it["meters"], price,
                     round(it["meters"] * price, 2), it["note"]))
                done["发货明细"] += 1

        for p in plan["payments"]:
            conn.execute(
                """INSERT INTO payment(customer_id, pay_date, amount, method,
                   ref_no, note) VALUES (?,?,?,?,?,?)""",
                (cust[p["customer"]], p["date"], p["amount"], p["method"],
                 p["ref_no"], p["note"]))
            done["收款"] += 1

    return done


def _lookup_price(conn, customer_id, item_id, process_id, on_date):
    """导入时在同一事务内取价：先按客户+面料+工艺，再退到通用价。"""
    if not process_id:
        return None
    row = conn.execute("SELECT fabric_id FROM inbound_item WHERE id=?",
                       (item_id,)).fetchone()
    fid = row["fabric_id"] if row else None
    for cond, args in ((" AND fabric_id=?", (fid,)) if fid else (None, None),
                       (" AND fabric_id IS NULL", ())):
        if cond is None:
            continue
        r = conn.execute(
            "SELECT unit_price FROM price WHERE customer_id=? AND process_id=?"
            + cond + " AND effective_date<=? ORDER BY effective_date DESC LIMIT 1",
            (customer_id, process_id) + args + (on_date,)).fetchone()
        if r:
            return r["unit_price"]
    return None

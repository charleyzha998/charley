# -*- coding: utf-8 -*-
"""老账本导入：把客户各自格式的 Excel 转成系统数据。

和 importer.py 的区别：importer.py 读的是我们自己定的模板（会计照着填），
这里读的是客户/我们自己历年手写的账本，每家格式都不一样，只能一家写一套规则。

对外只有两个入口：
    detect(path)            -> 猜这是哪家的什么表，返回 Rule 或 None
    preview(path, rule)     -> 解析成待导入数据 + 问题清单，不写库
    commit(preview_result)  -> 真正写库
"""

import datetime
import os
import re

from .importer import Report

# ---------------- 货源别名 ----------------
# 鹏川的货是从三个人手里拿的，账本上写的是简称。老板说货源要并进面料名称里，
# 比如「小陆」+「335尼龙斜纹」→「陆琴良335尼龙斜纹」。
SOURCE_ALIAS = {
    "小陆": "陆琴良",
    "陆琴良": "陆琴良",
    "陆秦良": "陆琴良",
    "邬建明": "邬剑明",
    "邬剑明": "邬剑明",
    "小邬": "邬剑明",
    "徐逸峰": "徐逸峰",
    "小徐": "徐逸峰",
}


def norm_source(name):
    """把账本上的货源简称还原成人名；不认识的原样返回。"""
    s = (name or "").strip()
    if not s:
        return ""
    return SOURCE_ALIAS.get(s, s)


def fabric_with_source(source, product):
    """货源 + 品名 → 面料名称。「小陆」「335尼龙斜纹」→「陆琴良335尼龙斜纹」。"""
    src = norm_source(source)
    p = (product or "").strip()
    if not p:
        return src
    if not src or p.startswith(src):
        return p
    return src + p


# ---------------- 支数（卷数）解析 ----------------
# 账本里的支数不一定是纯数字，见过：940+2大卷+3大卷 / 45 / 空 / 约50
_ROLL_NUM = re.compile(r"\d+(?:\.\d+)?")


def parse_rolls(v):
    """把「940+2大卷+3大卷」这种支数写法解析成卷数。

    返回 (卷数, 原文)。原文非纯数字时一并留着，写进备注，免得对不上账时查不到源头。
    """
    if v is None:
        return 0, ""
    if isinstance(v, (int, float)):
        return int(round(float(v))), ""
    s = str(v).strip()
    if not s:
        return 0, ""
    try:
        return int(round(float(s))), ""
    except ValueError:
        pass
    nums = _ROLL_NUM.findall(s)
    total = int(sum(float(n) for n in nums)) if nums else 0
    return total, s          # 保留原文


# ---------------- 日期 ----------------

def xl_date(v, datemode=0):
    """Excel 日期：序列号 / datetime / 「3/7来」这种手写文本，都尽量认出来。"""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date().strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        n = float(v)
        if n < 1 or n > 80000:       # 明显不是日期序列号
            return None
        try:
            import xlrd
            y, mo, d = xlrd.xldate_as_tuple(n, datemode)[:3]
            return datetime.date(y, mo, d).strftime("%Y-%m-%d")
        except Exception:
            base = datetime.date(1899, 12, 30)
            return (base + datetime.timedelta(days=int(n))).strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", s)
    if m:
        y, mo, d = (int(g) for g in m.groups())
        try:
            return datetime.date(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return None
    return None


def md_date(v, year):
    """「3/7来」「3.7」这种只有月日的手写日期，配上年份补全。"""
    if v is None:
        return None
    full = xl_date(v)
    if full:
        return full
    m = re.match(r"^(\d{1,2})[/.\-](\d{1,2})", str(v).strip())
    if not m:
        return None
    mo, d = int(m.group(1)), int(m.group(2))
    try:
        return datetime.date(year, mo, d).strftime("%Y-%m-%d")
    except ValueError:
        return None


# ---------------- 合并单元格补齐 ----------------

def fill_down(rows, keys):
    """账本里一个来货单位下面挂好几行颜色，只有第一行有值 —— 把空的往下补齐。

    rows 是 [dict, ...]，原地修改并返回。
    """
    last = {}
    for r in rows:
        for k in keys:
            v = r.get(k)
            if v is None or str(v).strip() == "":
                if k in last:
                    r[k] = last[k]
            else:
                last[k] = v
    return rows


def is_blank_row(r, keys):
    return all(r.get(k) is None or str(r.get(k)).strip() == "" for k in keys)


# ---------------- 金额校验 ----------------
# 老账本的「总金额」不导入 —— 系统里金额一律 米数 × 单价 自己算。
# 但拿它跟算出来的对一下很有用：差太多说明账本那行数字抄错了，导入前就能发现。

AMOUNT_TOL = 0.02        # 差 2% 以内算正常（手工账本常有抹零）


def check_amount(meters, unit_price, book_amount, tol=AMOUNT_TOL):
    """返回 None 表示对得上，否则返回一句给人看的提示。"""
    if not book_amount or not unit_price or not meters:
        return None
    try:
        calc = float(meters) * float(unit_price)
        book = float(book_amount)
    except (TypeError, ValueError):
        return None
    if book <= 0:
        return None
    diff = abs(calc - book)
    if diff <= max(1.0, book * tol):
        return None
    return (f"账本总金额 {book:,.2f}，按 {meters:g} 米 × {unit_price:g} 元算是 "
            f"{calc:,.2f}，差 {diff:,.2f} 元，请核对")


def check_group_amount(group_meters, unit_price, book_amount, tol=AMOUNT_TOL):
    """账本里单价和总金额只写在一组的第一行，金额是整组的小计。

    实测（鹏川 3 月）：(3633+3400) × 1.9 = 13362.70，(3042+3113) × 1.9 = 11694.50，
    都严丝合缝。所以拿整组米数去核金额，能反过来验证「这几行是不是同一组」——
    分组分错了，金额立刻对不上。金额本身不入库。
    """
    return check_amount(sum(group_meters or []), unit_price, book_amount, tol)


def is_total_row(v):
    """账本里夹着「合计：」小计行 —— 必须跳过，否则米数翻倍。

    实测（鹏川 3 月）：行 6~25 是明细，行 26 是「合计： 974 支 127504 米」，
    127504 × 0.85 = 108378.40 正好等于组首行写的总金额。
    """
    return v is not None and "合计" in str(v)


# ---------------- 统一读表（.xls 和 .xlsx 都能读）----------------

class Sheet:
    """把 openpyxl / xlrd 的差异包掉，统一成 cell(r, c) 一维接口（行列都从 1 起）。"""

    def __init__(self, name, grid, datemode=0):
        self.name = name
        self._g = grid                      # [[v, ...], ...]
        self.datemode = datemode
        self.nrows = len(grid)
        self.ncols = max((len(r) for r in grid), default=0)

    def cell(self, r, c):
        if 1 <= r <= self.nrows:
            row = self._g[r - 1]
            if 1 <= c <= len(row):
                return row[c - 1]
        return None

    def row(self, r, cols):
        return [self.cell(r, c) for c in cols]

    def date(self, r, c, year=None):
        v = self.cell(r, c)
        d = xl_date(v, self.datemode)
        if d is None and year:
            d = md_date(v, year)
        return d


def read_sheets(path, max_rows=2000, max_cols=64):
    """读出工作簿全部 sheet。账本常有几千列空列，这里按 max_cols 截断。"""
    ext = os.path.splitext(path)[1].lower()
    out = []
    if ext == ".xls":
        import xlrd
        bk = xlrd.open_workbook(path)
        for s in bk.sheets():
            grid = [[s.cell_value(r, c) for c in range(min(s.ncols, max_cols))]
                    for r in range(min(s.nrows, max_rows))]
            out.append(Sheet(s.name, grid, bk.datemode))
    else:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            grid = []
            for i, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_cols,
                                                 values_only=True)):
                grid.append(list(row))
            out.append(Sheet(ws.title, grid))
        wb.close()
    return out


def find_header(sheet, must_have, max_row=8):
    """找表头行：返回 (行号, {标题: 列号})。找不到返回 (None, {})。"""
    for r in range(1, min(sheet.nrows, max_row) + 1):
        cols = {}
        for c in range(1, sheet.ncols + 1):
            t = ("" if sheet.cell(r, c) is None else str(sheet.cell(r, c))).strip()
            if t and t not in cols:
                cols[t] = c
        if all(k in cols for k in must_have):
            return r, cols
    return None, {}


def sheet_month(name, default=None):
    """从表名取月份：「3月份」→3、「2026年4月」→4、「2026年5月)」→5。

    注意只能取月，取不了年 —— 逸峰的「1月」其实是次年 1 月。年份一律从数据里取。
    """
    m = re.search(r"(\d{1,2})\s*月", str(name).strip())
    return int(m.group(1)) if m else default


# ---------------- 解析结果 ----------------

class ParseResult:
    """解析产物。只是数据，不碰数据库。"""

    def __init__(self, customer, use_dye_lot=True):
        self.customer = customer
        self.use_dye_lot = use_dye_lot
        self.inbounds = []       # {in_date, dye_lot, fabric, color, rolls, meters, note}
        self.productions = []    # {done_date, fabric, color, process, rolls, meters,
                                 #  weight, ship_date, note}
        self.shipments = []      # {ship_date, fabric, process, meters, unit_price, note}
        self.prices = []         # {fabric, process, unit_price}
        self.payments = []
        # 期初欠款：老客户前面的账已结清，只结转一个总额（龚松权 3月底 316094.23）
        self.opening_debt = None
        self.opening_note = ""
        self.report = Report()

    @property
    def stats(self):
        return {"进仓": len(self.inbounds), "加工": len(self.productions),
                "发货": len(self.shipments), "价格": len(self.prices),
                "收款": len(self.payments)}

    def err(self, msg):
        self.report.errors.append(msg)

    def warn(self, msg):
        self.report.warnings.append(msg)


# ---------------- 账本里的杂事：拆工艺、洗米数、认小计 ----------------

# 品名/颜色里串进来的工艺。账本上「厚布杂布（贴60克黑针织）」是一格写完的，
# 品名归品名、工艺归工艺，系统里才好按「面料+工艺」查单价。
PROCESS_HINT = re.compile(
    r"(贴?\d+克(?:黑)?(?:针织|复合)|贴[白黑]膜|贴?低透明|贴?透明膜|贴?白膜|贴?黑膜|"
    r"做?PE膜|PE膜|防水|复合|单切边|双切边|定型)")

# 「大卷装」是包装规格：一个卷桩上直接卷很多米，不打成小卷。
# 不是品名也不是工艺 —— 拆出来写备注，卷数照样一卷算一卷。
PACK_HINT = re.compile(r"(大卷装|小卷装|大卷|小卷)")
CABINET_HINT = re.compile(r"(\d+号柜)")


def split_product(text):
    """把账本上一格写完的品名拆成 (品名, 工艺, 备注)。

    「厚布杂布（贴60克黑针织）」→ ("厚布杂布", "贴60克黑针织", "")
    「335尼龙斜/1号柜」        → ("335尼龙斜", "", "1号柜")
    「厚布杂布（大卷装）」      → ("厚布杂布", "", "大卷装")
    """
    s = ("" if text is None else str(text)).strip()
    if not s:
        return "", "", ""
    notes, proc = [], ""
    # 括号/斜杠里的附注逐个认领
    for part in re.split(r"[（(/]", s)[1:]:
        p = part.rstrip("）) ").strip()
        if not p:
            continue
        if not proc and PROCESS_HINT.fullmatch(p):
            proc = p
        elif PACK_HINT.fullmatch(p) or CABINET_HINT.fullmatch(p):
            notes.append(p)
        elif PROCESS_HINT.search(p):
            proc = proc or PROCESS_HINT.search(p).group(1)
        else:
            notes.append(p)
    name = re.split(r"[（(/]", s)[0].strip()
    return name, proc, "，".join(notes)


def color_or_process(text):
    """颜色列里也会串工艺（「贴70克黑针织」「60克黑针织」）。返回 (颜色, 工艺)。"""
    s = ("" if text is None else str(text)).strip()
    if not s:
        return "", ""
    if s.startswith("贴") or "克黑针织" in s or PROCESS_HINT.fullmatch(s):
        return "", s
    color, _proc, note = split_product(s)      # 「杂色（大卷）」
    return (color + ("（" + note + "）" if note else "")), ""


# 米数写成「10000(实际米数：9800米）」—— 括号里的才是实际数
_PAREN_METERS = re.compile(r"[（(].*?(\d+(?:\.\d+)?)\s*米?\s*[）)]")


def parse_meters(v):
    """返回 (米数, 原文)。原文非纯数字时留着写备注。"""
    if v is None:
        return 0.0, ""
    if isinstance(v, (int, float)):
        return float(v), ""
    s = str(v).strip()
    if not s:
        return 0.0, ""
    try:
        return float(s), ""
    except ValueError:
        pass
    m = _PAREN_METERS.search(s)
    if m:                                   # 括号内是实际米数，以它为准
        return float(m.group(1)), s
    nums = _ROLL_NUM.findall(s)
    return (float(nums[0]) if nums else 0.0), s


# 汇总/小计行的标记词。这些行的数字是上面几行的和，导进去米数就翻倍。
_SUM_WORDS = ("合计", "总计", "小计", "上月结欠", "累计结欠", "本月加工费",
              "本月实收", "月底结欠", "总发货", "月总打款", "加工费")
_TRIP = re.compile(r"^第[一二三四五六七八九十\d]+趟")


def is_sum_text(v):
    """是不是汇总行的标记（含鹏川发货的「第一趟共：」「总计：」）。"""
    if v is None:
        return False
    s = str(v).strip()
    if not s:
        return False
    if _TRIP.match(s):
        return True
    return any(w in s for w in _SUM_WORDS)


def is_sum_row(sheet, r, cols):
    """cols 里任意一格是汇总标记，整行就当汇总跳掉。"""
    return any(is_sum_text(sheet.cell(r, c)) for c in cols)


# ---------------- 规则框架 ----------------
# 每家账本一套规则。认表靠文件名 + 表头特征；解析只产数据，不碰数据库。

class Rule:
    name = ""            # 「鹏川纺织·入库明细」
    customer = ""        # 建哪个客户
    use_dye_lot = True   # 这家要不要按缸号管库存
    filename_hint = ()   # 文件名里得同时含这些词

    def match(self, path, sheets):
        base = os.path.basename(path)
        return all(h in base for h in self.filename_hint)

    def parse(self, sheets, year=None):
        raise NotImplementedError

    # -- 子类常用的小工具 --
    @staticmethod
    def month_sheets(sheets):
        """只要「3月份」「2026年4月」这种月份表，跳过 Sheet1/单价/空表。"""
        out = []
        for s in sheets:
            if re.search(r"\d+\s*月", s.name) and s.nrows > 1:
                out.append(s)
        return out

    @staticmethod
    def _take_payment(s, r, c_date, c_amt, yr, res, fallback_col=None,
                      fallback_date=None):
        """打款是稀疏列：只有发生打款那行才填。抽成独立收款记录，
        别当成该发货行的属性 —— 否则一笔货款会挂到某笔货上，对账就乱了。

        账本上常常只写打款金额、不写打款日期（逸峰 6 处、鹏川若干）。
        钱是真的，绝不能丢：日期依次退让 —— 打款日期 → 同行的业务日期 →
        该月 1 号，每退一步都提醒人去核对。
        """
        if not c_date or not c_amt:
            return
        amt = _num(s.cell(r, c_amt))
        if not amt or amt <= 0:
            return
        d = s.date(r, c_date, yr)
        note = ""
        if not d and fallback_col:
            d = s.date(r, fallback_col, yr)
        if not d:
            d = fallback_date
        if not d:
            res.warn("[%s] 第%d行 有打款 %s 元但认不出日期，没导入，请人工补"
                     % (s.name, r, _g(amt)))
            return
        if not s.date(r, c_date, yr):
            note = "账本没写打款日期，日期是估的"
            res.warn("[%s] 第%d行 打款 %s 元没写日期，暂记为 %s，请核对"
                     % (s.name, r, _g(amt), d))
        res.payments.append({"pay_date": d, "amount": amt, "method": "转账",
                             "sheet": s.name, "row": r, "note": note})


def book_year(sheets, date_cols):
    """整本账的年份：哪张月表里能读到日期就用它的年。

    有的月份一个日期都没写（鹏川对账 4月份），但同一个文件别的月表有 ——
    一本账不会跨太多年，借过来当兜底比留空强。
    """
    for s in sheets:
        y = year_of(s, date_cols)
        if y:
            return y
    return None


def _has_data(sheet, num_cols):
    """这张表除了表头，到底有没有数？空表是提前建好的月表，不算问题。"""
    for r in range(2, sheet.nrows + 1):
        for c in num_cols:
            if _num(sheet.cell(r, c)):
                return True
    return False


def year_of(sheet, date_cols, default=None):
    """年份从数据里取 —— sheet 名不可信（「1月」其实是次年 1 月）。"""
    for r in range(1, min(sheet.nrows, 60) + 1):
        for c in date_cols:
            d = xl_date(sheet.cell(r, c), sheet.datemode)
            if d:
                return int(d[:4])
    return default


# ---------------- 鹏川纺织 ----------------
# 老板说的：左边是进仓（客户给我的坯布），右边是我加工好、在我厂里的成品。
# 所以鹏川要管库存，三段都有。两块行不对齐，各读各的。
#
# 账本没有缸号 —— 用「来货日期+货源+品名+颜色」当批次标识，见 lot_key()。

PC_IN_COLS = {"date": 1, "party": 2, "product": 3, "color": 4,
              "rolls": 5, "meters": 6, "sum": 7}
PC_FIN_COLS = {"date": 9, "party": 10, "product": 11, "color": 12,
               "rolls": 13, "meters": 14, "weight": 15, "ship": 16,
               "price": 17, "amount": 18, "loss": 19, "note": 20}


def lot_key(in_date, source, fabric, color, seq=None):
    """没缸号的客户，用这个当缸号：`0314-陆琴良335尼龙斜-黑色`。

    缸号在系统里同客户内唯一，所以重名时补序号。
    """
    d = (in_date or "")[5:].replace("-", "")
    parts = [p for p in (d, (norm_source(source) or "") + (fabric or ""), color or "") if p]
    k = "-".join(parts) or "无编号"
    return k if seq in (None, 0) else "%s#%d" % (k, seq + 1)


class PengchuanInboundRule(Rule):
    """鹏川·入库明细：左块 → 进仓坯布，右块 → 加工成品。

    实测结构（3月份~10月份 8 张月表 + 黑针织发货明细）：
      第 1 行表头；左块 1-7 列（来货日期/来货单位/品名/颜色/来货支数/来货米数/合计）；
      第 8 列是 ✅️ 标记；右块 9-20 列（成品）。两块**行不对齐**，必须各读各的。
      来货单位/品名只写在一组第一行，往下补齐。
      单价/总金额是整组小计 —— 只用来校验，不入库（老板：成品块总金额可以不用记录）。
    """
    name = "鹏川纺织·入库明细"
    customer = "鹏川纺织"
    filename_hint = ("鹏川", "入库")

    def _parse_all(self, sheets, year=None):
        res = ParseResult(self.customer, use_dye_lot=True)
        seen = {}
        for s in self.month_sheets(sheets):
            yr = year_of(s, [PC_IN_COLS["date"], PC_FIN_COLS["ship"]], year) or year
            if not yr:
                # 9月/10月是提前建好的空表，只有表头没有数据 —— 那是正常的，
                # 别报警吓人。真有数据却认不出年份才值得提醒。
                if _has_data(s, (PC_IN_COLS["meters"], PC_FIN_COLS["meters"])):
                    res.warn("[%s] 认不出年份，整表跳过" % s.name)
                continue
            self._parse_inbound(s, yr, res, seen)
            self._parse_finished(s, yr, res)
        return res

    # ---- 左块：来货 → 进仓 ----
    # 分组：一条来货记录挂多行颜色，只有首行写来货日期/单位/品名，
    # 首行的「合计（米）」是整组米数 —— 拿它验分组对不对。
    def _parse_inbound(self, s, yr, res, seen):
        C = PC_IN_COLS
        rows = []
        for r in range(2, s.nrows + 1):
            if is_sum_row(s, r, [C["party"], C["product"], C["color"]]):
                continue
            rows.append({
                "r": r,
                "date": s.cell(r, C["date"]), "party": s.cell(r, C["party"]),
                "product": s.cell(r, C["product"]), "color": s.cell(r, C["color"]),
                "rolls": s.cell(r, C["rolls"]), "meters": s.cell(r, C["meters"]),
                "sum": s.cell(r, C["sum"]),
            })
        # 合并单元格：来货日期/单位/品名只写在组首行
        fill_down(rows, ["date", "party", "product"])
        group, book_sum, head = [], None, None
        for row in rows:
            if row["sum"] not in (None, ""):        # 新组开始
                self._check_sum(s, res, "来货", head, group, book_sum)
                group, book_sum, head = [], _num(row["sum"]), row["r"]
            if is_blank_row(row, ["meters", "rolls"]):
                continue
            meters, mraw = parse_meters(row["meters"])
            rolls, rraw = parse_rolls(row["rolls"])
            if meters <= 0 and rolls <= 0:
                continue
            group.append(meters)
            d = xl_date(row["date"], s.datemode) or md_date(row["date"], yr)
            if not d:
                res.warn("[%s] 第%d行 来货日期「%s」认不出，按当月1号记"
                         % (s.name, row["r"], _short(row["date"])))
                d = "%04d-%02d-01" % (yr, sheet_month(s.name, 1))
            product, proc, pnote = split_product(row["product"])
            color, cproc = color_or_process(row["color"])
            base = lot_key(d, row["party"], product, color)
            seq = seen.get(base, 0)
            seen[base] = seq + 1
            notes = [x for x in (pnote, mraw and "米数原文：" + mraw,
                                 rraw and "支数原文：" + rraw) if x]
            res.inbounds.append({
                "in_date": d, "dye_lot": lot_key(d, row["party"], product, color, seq),
                "fabric": fabric_with_source(row["party"], product),
                "color": color, "rolls": rolls, "meters": meters,
                "process": proc or cproc, "sheet": s.name, "row": row["r"],
                "note": "；".join(notes)})
        self._check_sum(s, res, "来货", head, group, book_sum)

    # ---- 右块：成品 → 加工 ----
    # 分组（实测 3/4/5/7/8 月都对得上）：一组是连续的若干行，
    #   开头写单价和总金额（金额是整组小计），结尾可能有「合计：」行。
    # 组的分界有三种，都要认：
    #   ① 「合计：」行 —— 组的结尾，它的米数/支数正好是整组的和，拿来验分组
    #   ② 空行 —— 8月份好几组之间隔着空行，没写合计
    #   ③ 又出现单价 —— 说明换了个价，是新的一组
    # 注意组内可以有好几个「来货日期」（一个合计能盖住两批来货），
    # 所以来货日期不是分界，别拿它切组。
    def _parse_finished(self, s, yr, res):
        C = PC_FIN_COLS
        rows = []
        for r in range(2, s.nrows + 1):
            row = {"r": r, "sum_row": is_sum_row(s, r, [C["party"], C["product"],
                                                        C["color"]])}
            for k in ("date", "party", "product", "color", "rolls", "meters",
                      "weight", "ship", "price", "amount", "note"):
                row[k] = s.cell(r, C[k])
            # 空行要在补齐之前判，补齐会把日期/单位填进来
            row["blank"] = is_blank_row(row, ["date", "party", "product", "color",
                                              "rolls", "meters", "ship", "price"])
            rows.append(row)
        # 颜色每行都自己写，不补；日期/单位/品名才补
        fill_down([r for r in rows if not r["sum_row"]], ["date", "party", "product"])
        group, price, book, head = [], None, None, None

        def close(book_meters=None):
            self._check_sum(s, res, "成品", head, group, book_meters)
            self._check_money(s, res, group, price, book)

        for row in rows:
            if row["sum_row"]:                       # ①「合计：」= 组的结尾
                bk_meters, _ = parse_meters(row["meters"])
                close(bk_meters or None)
                group, price, book, head = [], None, None, None
                continue
            if row["blank"] and group:               # ② 空行隔开
                close()
                group, price, book, head = [], None, None, None
                continue
            if row["price"] not in (None, "") and group:      # ③ 换价 = 换组
                close()
                group, price, book, head = [], None, None, None
            if row["price"] not in (None, ""):
                price, book, head = _num(row["price"]), _num(row["amount"]), row["r"]
            if head is None:
                head = row["r"]
            if is_blank_row(row, ["meters", "rolls"]):
                continue
            meters, mraw = parse_meters(row["meters"])
            rolls, rraw = parse_rolls(row["rolls"])
            if meters <= 0 and rolls <= 0:
                continue
            d = xl_date(row["date"], s.datemode) or md_date(row["date"], yr)
            ship = xl_date(row["ship"], s.datemode) or md_date(row["ship"], yr)
            if not d:
                d = ship        # 加工日期常空（「3/7来」是来货日），用发货日兜底
            if not d:
                res.warn("[%s] 第%d行 成品没有日期，跳过" % (s.name, row["r"]))
                continue
            product, proc, pnote = split_product(row["product"])
            color, cproc = color_or_process(row["color"])
            weight, _wraw = parse_meters(row["weight"])
            notes = [x for x in (pnote, _short(row["note"]),
                                 mraw and "米数原文：" + mraw,
                                 rraw and "支数原文：" + rraw) if x]
            res.productions.append({
                "done_date": d, "ship_date": ship,
                "fabric": fabric_with_source(row["party"], product),
                "color": color, "process": proc or cproc,
                "rolls": rolls, "meters": meters, "weight": weight or None,
                "unit_price": price or 0, "sheet": s.name, "row": row["r"],
                "note": "；".join(notes)})
            group.append(meters)
        close()

    def parse(self, sheets, year=None):
        res = self._parse_all(sheets, year)
        pc_link(res)          # 加工挂到坯布批次上，库存才扣得下来
        return res

    @staticmethod
    def _check_sum(s, res, what, head, group, book_sum):
        """账本自己写了组小计米数 —— 和我们解析出来的对一下，分组错了立刻现形。"""
        if not group or not book_sum:
            return
        got = sum(group)
        if abs(got - book_sum) > max(1.0, book_sum * 0.005):
            res.warn("[%s] 第%s行起的%s组：账本小计 %s 米，逐行加起来是 %s 米，请核对"
                     % (s.name, head, what, _g(book_sum), _g(got)))

    @staticmethod
    def _check_money(s, res, group, price, book):
        """账本总金额不入库，只拿来验分组：对不上说明这几行没归对组。"""
        if not group or not price or not book:
            return
        msg = check_group_amount(group, price, book)
        if msg:
            res.warn("[%s] %s" % (s.name, msg))


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def pc_link(res):
    """鹏川：把加工挂到进仓批次上 —— 不然进仓的米数永远挂在库存里下不来。

    账本左右两块行不对齐，颜色写法还不统一（左边写「杂色」右边空着），
    所以逐缸对不上。能靠得住的只有品名：**同一个品名，按来货先后（先进先出）
    把加工的米数从坯布上扣掉**。这也是车间实际的做法 —— 同一个品名的坯布堆在
    一起，先来的先用。

    配不上的（加工里有、进仓里没有的品名，实测 27 种 65 万米）说明那批坯布是
    账本开始记之前进的，就不挂缸号，只当成品记上，不去扣不存在的坯布。
    """
    pool = {}
    for x in sorted(res.inbounds, key=lambda x: x["in_date"]):
        pool.setdefault(x["fabric"], []).append({"lot": x["dye_lot"], "left": x["meters"]})

    linked = unlinked = 0.0
    for p in sorted(res.productions, key=lambda x: x["done_date"]):
        need = p["meters"]
        got = []
        for b in pool.get(p["fabric"], []):
            if need <= 0.01:
                break
            if b["left"] <= 0.01:
                continue
            take = min(b["left"], need)
            b["left"] -= take
            need -= take
            got.append((b["lot"], take))
        if got:
            # 一笔加工可能吃掉好几缸坯布 —— 系统里一条加工只能挂一缸，
            # 挂吃得最多的那缸，其余写进备注，人查得到。
            got.sort(key=lambda t: -t[1])
            p["dye_lot"] = got[0][0]
            if len(got) > 1:
                extra = "、".join("%s %s米" % (l, _g(m)) for l, m in got[1:])
                p["note"] = "；".join(x for x in (p.get("note"), "另用坯布：" + extra) if x)
            linked += p["meters"] - need
        if need > 0.01:
            unlinked += need
    if unlinked > 1:
        res.warn("有 %s 米加工在进仓记录里找不到对应的坯布（品名对不上或坯布是"
                 "账本记录之前进的），这部分只记成品、不扣坯布库存，请核对。"
                 % _g(unlinked))
    return linked, unlinked


def pc_link_ship(in_res, ship_res):
    """鹏川：把对账明细里的发货挂到入库明细的成品上。

    两个文件是分开记的，**品名写法还不一样**：入库明细带货源（「盛泽阿提335尼龙斜」），
    对账明细不带（「335尼龙斜」）—— 所以按品名配一个都配不上。

    真正靠得住的是「发货日期 + 米数」：入库明细右边那块本来就写了发货日期，
    米数跟对账明细一模一样（实测 3633 米那笔两边完全对上）。所以先按
    「发货日期+米数」精确配，配不上的再退一步按米数配（日期差几天的情况）。
    配不上的照样进对账单（钱是真的），只是不扣成品库存。

    发货记录里写上 `prod_key`，commit 时换成真正的 production_id。
    """
    for i, p in enumerate(in_res.productions):
        p.setdefault("key", "P%d" % i)

    # 索引一：发货日期+米数（四舍五入到米，账本偶尔有小数）
    exact, loose = {}, {}
    for p in in_res.productions:
        m = round(p.get("meters") or 0)
        if p.get("ship_date"):
            exact.setdefault((p["ship_date"], m), []).append(p)
        loose.setdefault(m, []).append(p)

    unlinked = 0.0
    for sh in sorted(ship_res.shipments, key=lambda x: x.get("ship_date") or ""):
        m = round(sh.get("meters") or 0)
        hit = None
        for p in exact.get((sh.get("ship_date"), m), []):
            if not p.get("_used"):
                hit = p
                break
        if hit is None:
            for p in loose.get(m, []):
                if not p.get("_used"):
                    hit = p
                    break
        if hit is not None:
            hit["_used"] = True
            sh["prod_key"] = hit["key"]
        else:
            unlinked += sh.get("meters") or 0
    if unlinked > 1:
        ship_res.warn("有 %s 米发货在加工记录里找不到对应的成品 —— 应该是账本"
                      "开始记之前做好的货。这些照样进对账单，但不扣成品库存。"
                      % _g(unlinked))
    return unlinked


class PengchuanStatementRule(Rule):
    """鹏川·对账明细：一行一笔发货，另有「单价」页是真正的价目表。

    实测结构（3月份~9月份 + 单价）：
      表头在第 1 行，两个版本 ——
        3/4 月：日期|产品名称|加工要求|数量（米）|单价（元）|金额|备注
        5-9 月：发货日期|产品名称|加工要求|支数|数量（米）|单价（元）|金额|
                打款日期|金额|累计结欠金额|备注     （「金额」出现两次！）
      所以列一律按标题找，不能按固定列号。「金额」重名 —— 靠打款日期定位第二个。
      表尾有「合计 / 上月结欠余额 / 本月加工费 / 本月实收金额」等汇总行，跳过。
      打款是稀疏列：只有发生打款那行才填，单独抽成收款记录，不是该发货行的属性。
    """
    name = "鹏川纺织·对账明细"
    customer = "鹏川纺织"
    filename_hint = ("鹏川", "对账")

    def parse(self, sheets, year=None):
        res = ParseResult(self.customer, use_dye_lot=True)
        for s in sheets:
            if s.name.strip() == "单价":
                self._parse_prices(s, res)
        months = self.month_sheets(sheets)
        # 4月份整月一个日期都没写 —— 年份得从别的月表借。
        # 不借的话那 58 行只能报「暂记为 空」，等于白提醒。
        year = year or book_year(months, [1, 2])
        for s in months:
            self._parse_month(s, res, year)
        return res

    # ---- 「单价」页 → 价目表 ----
    def _parse_prices(self, s, res):
        hdr, cols = find_header(s, ["产品名称", "加工要求", "单价"], max_row=6)
        if not hdr:
            res.warn("[单价] 找不到表头（产品名称/加工要求/单价），价目表没导")
            return
        for r in range(hdr + 1, s.nrows + 1):
            fabric = _short(s.cell(r, cols["产品名称"]), 60)
            proc = _short(s.cell(r, cols["加工要求"]), 60)
            price = _num(s.cell(r, cols["单价"]))
            if not fabric or not price:
                continue
            res.prices.append({"fabric": fabric, "process": proc, "unit_price": price})

    # ---- 月份表 → 发货 + 收款 ----
    def _parse_month(self, s, res, year):
        hdr, cols = find_header(s, ["产品名称", "加工要求"], max_row=4)
        if not hdr:
            res.warn("[%s] 找不到表头，整表跳过" % s.name)
            return
        c_date = cols.get("发货日期") or cols.get("日期")
        c_meters = cols.get("数量（米）") or cols.get("数量(米)")
        c_price = cols.get("单价（元）") or cols.get("单价(元)") or cols.get("单价")
        c_rolls = cols.get("支数")
        c_note = cols.get("备注")
        c_pay_date = cols.get("打款日期")
        # 「金额」这个标题出现两次：前一个是加工金额，打款日期后面那个是打款金额。
        # find_header 只留第一个，第二个按位置取（紧挨打款日期右边）。
        c_amount = cols.get("金额")
        c_pay_amt = (c_pay_date + 1) if c_pay_date else None

        yr = year_of(s, [c for c in (c_date,) if c], year) or year
        for r in range(hdr + 1, s.nrows + 1):
            fabric = _short(s.cell(r, cols["产品名称"]), 60)
            # 汇总行：整行任意一格写着「合计/上月结欠/本月加工费/总打款…」都算。
            # 表尾那行长这样：|||363244|合计|414040||552145|183875 ——
            # 「合计」在单价列、打款金额在打款列，只看头几列会漏。
            if any(is_sum_text(s.cell(r, c)) for c in range(1, s.ncols + 1)):
                continue
            if _num(fabric) is not None:      # 汇总行溢出来的数字
                continue
            self._take_payment(s, r, c_pay_date, c_pay_amt, yr, res,
                               fallback_col=c_date)
            meters, mraw = parse_meters(s.cell(r, c_meters)) if c_meters else (0.0, "")
            if meters <= 0:
                continue
            if not fabric:
                # 4月份整月只写了米数/单价/金额，产品名和日期全空 —— 钱是真的
                # （账本合计 461165 米 / 467587.95 元），不能悄悄丢。
                # 照样导，面料名写「未注明」，日期用当月，让人在软件里补。
                fabric = "未注明"
                res.warn("[%s] 第%d行 账本没写产品名称，先按「未注明」导入，"
                         "请在软件里补上" % (s.name, r))
            rolls, rraw = parse_rolls(s.cell(r, c_rolls)) if c_rolls else (0, "")
            d = s.date(r, c_date, yr) if c_date else None
            if not d:
                mo = sheet_month(s.name)      # 「4月份」→ 4
                d = "%04d-%02d-01" % (yr, mo) if (yr and mo) else None
                res.warn("[%s] 第%d行 账本没写发货日期，暂记为 %s，请核对"
                         % (s.name, r, d or "空"))
            price = _num(s.cell(r, c_price)) if c_price else None
            book = _num(s.cell(r, c_amount)) if c_amount else None
            msg = check_amount(meters, price, book)
            if msg:
                res.warn("[%s] 第%d行 %s" % (s.name, r, msg))
            product, proc, pnote = split_product(fabric)
            notes = [x for x in (pnote, _short(s.cell(r, c_note)) if c_note else "",
                                 mraw and "米数原文：" + mraw,
                                 rraw and "支数原文：" + rraw) if x]
            res.shipments.append({
                "ship_date": d, "fabric": product,
                "process": proc or _short(s.cell(r, cols["加工要求"]), 40),
                "rolls": rolls, "meters": meters, "unit_price": price or 0,
                "sheet": s.name, "row": r, "note": "；".join(notes)})


# ---------------- 龚松权 ----------------
# 老板说：龚松权前面的账都结完了，库存也消掉了，只记一个总欠款金额；
# 只有 22、23 年会剩一点库存。发货对账单是从 2026 年 4 月开始记的。
#
# **缸号会重复用**（老板确认：「他有些缸号数字是一样的」）。实测缸号 4818
# 在 2026-03 和 2026-07 各来过一批，米数完全不同。所以缸号不能单独当批次
# 标识，一律用「日期+缸号」认批次 —— 系统里缸号要唯一，就写成 `4818@0702`。

GS_CUTOFF = "2026-04-01"       # 发货对账单的起点：这天之前的账已结清


def gs_lot(v):
    """缸号规范化：数字 3778.0 → 「3778」，文本 '0264' 保持补零写法。

    账本里同一个缸号一半写成数字、一半写成 4 位文本，不统一就配不上货。
    """
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return "%04d" % int(v) if 0 < v < 10000 else "%g" % v
    if isinstance(v, int):
        return "%04d" % v if 0 < v < 10000 else str(v)
    s = str(v).strip()
    if s.isdigit() and len(s) <= 4:
        return "%04d" % int(s)
    return s


def gs_batch_no(lot, date):
    """批次标识 = 缸号@月日。缸号重复用，光缸号认不出是哪一批。"""
    if not lot:
        return "无缸号"
    d = (date or "").replace("-", "")[4:]
    return ("%s@%s" % (lot, d)) if d else lot


class GongsongquanInboundRule(Rule):
    """龚松权入库及库存明细：一行一缸，四年流水（2022-07 ~ 2026-08，371 行）。

    列：日期|缸号|面料名称|颜色|进仓卷数|进仓米数|库存|发完✅️|备注|(标记)

    「库存」是老板手工维护的当前剩余米数，不是原始单据，所以**不直接导**，
    只拿来推算期初：
      · 2026-04-01 之后进的货 —— 照原样导成进仓，发货由对账单自己扣，不碰库存列
      · 之前的老货 —— 账已结清、库存已消（老板确认），只有还剩东西的才导，
        期初 = 现库存 + 4月起发掉的（这部分要留着给对账单扣），发完的整批跳过
    「成品待发」那张表是老板临时记的待发清单，同样的缸号在主表里都有，不导。
    """
    name = "龚松权·入库及库存明细"
    customer = "龚松权"
    filename_hint = ("龚松权", "入库")

    GS_COLS = ["日期", "缸号", "面料名称", "颜色", "进仓卷数", "进仓米数", "库存"]

    def parse(self, sheets, year=None, sent_by_lot=None):
        """sent_by_lot: {(缸号, 进仓日期): 4月起已发米数}，由发货规则算好传进来。

        单独跑（没传）时按「老货全部按现库存导」处理，米数会偏小 —— 所以
        导入向导里两张表要一起导，见 GongsongquanShipmentRule.sent_by_lot()。
        """
        res = ParseResult(self.customer)
        sent = sent_by_lot or {}
        for s in sheets:
            hdr, cols = find_header(s, ["缸号", "进仓米数"], max_row=4)
            if not hdr:
                continue          # 「成品待发」没表头，跳过
            self._parse(s, hdr, cols, res, sent)
        return res

    def read_batches(self, sheets):
        """只读原始行（不做期初处理），给 gs_allocate 配货用。"""
        out = []
        for s in sheets:
            hdr, cols = find_header(s, ["缸号", "进仓米数"], max_row=4)
            if not hdr:
                continue
            yr = year_of(s, [cols["日期"]], 2026)
            for r in range(hdr + 1, s.nrows + 1):
                lot = gs_lot(s.cell(r, cols["缸号"]))
                meters, _ = parse_meters(s.cell(r, cols["进仓米数"]))
                d = s.date(r, cols["日期"], yr)
                if meters > 0 and d:
                    out.append({"lot_raw": lot, "in_date": d, "meters": meters, "row": r})
        return out

    def _parse(self, s, hdr, cols, res, sent):
        c_date, c_lot = cols["日期"], cols["缸号"]
        c_fab, c_color = cols["面料名称"], cols.get("颜色")
        c_rolls, c_meters = cols.get("进仓卷数"), cols["进仓米数"]
        c_stock, c_note = cols.get("库存"), cols.get("备注")
        yr = year_of(s, [c_date], 2026)
        seen = {}
        for r in range(hdr + 1, s.nrows + 1):
            lot = gs_lot(s.cell(r, c_lot))
            fab = _short(s.cell(r, c_fab), 60)
            meters, mraw = parse_meters(s.cell(r, c_meters))
            if not (lot or fab) or meters <= 0:
                continue          # 表尾几十行只有一个 0，跳掉
            d = s.date(r, c_date, yr)
            if not d:
                res.warn("[%s] 第%d行 缸号%s 没有进仓日期，这批先跳过，请人工补"
                         % (s.name, r, lot or "?"))
                continue
            stock = _num(s.cell(r, c_stock)) if c_stock else None
            note = _short(s.cell(r, c_note), 60) if c_note else ""

            if d < GS_CUTOFF:
                # 老货：账已结、库存已消，只导「现在还剩的 + 4月起发掉的」
                shipped = sent.get((lot, d), 0.0)
                keep = max(0.0, stock or 0) + shipped
                if keep <= 0:
                    continue      # 早就发完了，不用进系统
                if keep > meters + 1:
                    res.warn("[%s] 第%d行 缸号%s 期初推算 %s 米超过当初进仓 %s 米，"
                             "按进仓米数导，请核对" % (s.name, r, lot, _g(keep), _g(meters)))
                    keep = meters
                rolls = 0
                if meters > 0 and c_rolls:
                    orig = parse_rolls(s.cell(r, c_rolls))[0]
                    rolls = int(round(orig * keep / meters)) if orig else 0
                note = "；".join(x for x in (
                    "期初结转（原进仓 %s 米，%s 进的货）" % (_g(meters), d), note) if x)
                meters = keep
            else:
                rolls, rraw = parse_rolls(s.cell(r, c_rolls)) if c_rolls else (0, "")
                if rraw:
                    note = "；".join(x for x in (note, "卷数原文：" + rraw) if x)
                if mraw:
                    note = "；".join(x for x in (note, "米数原文：" + mraw) if x)

            # 同一天同一缸号来两批（有过），补序号免得缸号撞车
            key = gs_batch_no(lot, d)
            seen[key] = seen.get(key, 0) + 1
            if seen[key] > 1:
                key = "%s#%d" % (key, seen[key])
            res.inbounds.append({
                "in_date": d, "dye_lot": key, "fabric": fab,
                "color": _short(s.cell(r, c_color), 40) if c_color else "",
                "rolls": rolls, "meters": meters,
                "sheet": s.name, "row": r, "lot_raw": lot, "note": note})


class GongsongquanShipmentRule(Rule):
    """龚松权发货对账单：月表 + 一张「单价」表。格式很规整，是三家里最干净的。

    表头两版，只是缸号和发货日期换了个位置（所以按标题名找列）：
      4月份    ：缸号|发货日期|面料名称|颜色|发货卷数|发货米数|加工方式|单价|金额|打款日期|客户打款|累计结欠金额
      5-9月份  ：发货日期|缸号|…（其余相同，7月「客户打款」写成「客户打款金额」）

    表尾三行是账，不是发货：
      · 「X月底累计结欠金额」——上月结转，只有 4 月那张要拿（期初欠款）
      · 「总计」行 —— 拿来核对解析结果
      · 「X月加工费 / X月总打款 / 累计结欠金额」——表头式的标签行
    """
    name = "龚松权·发货对账单"
    customer = "龚松权"
    filename_hint = ("龚松权", "发货")

    def parse(self, sheets, year=None):
        res = ParseResult(self.customer)
        for s in sheets:
            if s.name == "单价":
                self._parse_price(s, res)
                continue
            hdr, cols = find_header(s, ["缸号", "发货米数"], max_row=4)
            if not hdr:
                continue
            self._parse_month(s, hdr, cols, res, year)
        return res

    def _parse_price(self, s, res):
        hdr, cols = find_header(s, ["面料名称", "单价"], max_row=4)
        if not hdr:
            return
        for r in range(hdr + 1, s.nrows + 1):
            fab = _short(s.cell(r, cols["面料名称"]), 60)
            price = _num(s.cell(r, cols["单价"]))
            if not fab or not price:
                continue          # 有几行只写了面料名没写价，跳过
            res.prices.append({
                "fabric": fab,
                "process": _short(s.cell(r, cols.get("加工方式", 0)), 40)
                           if "加工方式" in cols else "",
                "unit_price": price, "sheet": s.name, "row": r})

    def _parse_month(self, s, hdr, cols, res, year):
        c_date, c_lot = cols["发货日期"], cols["缸号"]
        c_fab, c_color = cols["面料名称"], cols.get("颜色")
        c_rolls, c_meters = cols.get("发货卷数"), cols["发货米数"]
        c_proc, c_price = cols.get("加工方式"), cols.get("单价")
        c_amount = cols.get("金额")
        c_pay_date = cols.get("打款日期")
        c_pay_amt = cols.get("客户打款") or cols.get("客户打款金额")
        c_owe = cols.get("累计结欠金额")
        yr = year_of(s, [c_date], year) or year
        mo = sheet_month(s.name)
        fb = ("%04d-%02d-28" % (yr, mo)) if (yr and mo) else None

        for r in range(hdr + 1, s.nrows + 1):
            # 表尾「总计」行、「X月底累计结欠金额」标签行要跳过。
            # 但不能整行扫「合计」二字 —— 4月 r77 是真发货（2869缸 342米），
            # 老板顺手在那一行右边写了「3月底累计结欠金额 316094.23」，
            # 整行扫会把这笔货误杀。只认「总计」这一列自己写的标记。
            mark = "".join(str(s.cell(r, c) or "") for c in (c_fab, c_color, c_lot))
            self._take_opening(s, r, c_owe, res)
            if is_sum_text(mark) or "结欠" in mark:
                continue
            self._take_payment(s, r, c_pay_date, c_pay_amt, yr, res,
                               fallback_col=c_date, fallback_date=fb)

            meters, mraw = parse_meters(s.cell(r, c_meters))
            if meters <= 0:
                continue
            fab = _short(s.cell(r, c_fab), 60)
            lot = gs_lot(s.cell(r, c_lot))
            d = s.date(r, c_date, yr)
            if not d:
                d = fb
                res.warn("[%s] 第%d行 缸号%s 没写发货日期，暂记为 %s，请核对"
                         % (s.name, r, lot or "?", d or "空"))
            price = _num(s.cell(r, c_price)) if c_price else None
            msg = check_amount(meters, price, _num(s.cell(r, c_amount)) if c_amount else None)
            if msg:
                res.warn("[%s] 第%d行 缸号%s %s" % (s.name, r, lot or "?", msg))
            rolls, rraw = parse_rolls(s.cell(r, c_rolls)) if c_rolls else (0, "")
            notes = [x for x in (mraw and "米数原文：" + mraw,
                                 rraw and "卷数原文：" + rraw) if x]
            res.shipments.append({
                "ship_date": d, "dye_lot": lot, "fabric": fab,
                "color": _short(s.cell(r, c_color), 40) if c_color else "",
                "process": _short(s.cell(r, c_proc), 40) if c_proc else "",
                "rolls": rolls, "meters": meters, "unit_price": price or 0,
                "sheet": s.name, "row": r, "note": "；".join(notes)})

        self._check_total(s, cols, res)

    @staticmethod
    def _check_total(s, cols, res):
        """拿账本表尾的「总计」核对本月明细。

        实测 5~8 月的总计格都是 =SUM 公式，分毫不差；只有 4 月的卷数/米数是
        **手打的**（842 卷 / 134791 米），比明细多 16 卷 / 2843 米 —— 金额格
        是公式，反而对得上。所以明细是准的，那两个数是老板抄错了，提醒一声就行。
        """
        c_m, c_r = cols["发货米数"], cols.get("发货卷数")
        for r in range(s.nrows, 1, -1):
            # 从下往上找，但要认准写着「总计」且那一行确实有数的行 ——
            # 它下面还有「4月加工费 / 4月总打款」标签行，也含小计字样但没数。
            row = [s.cell(r, c) for c in range(1, s.ncols + 1)]
            if not any(v is not None and "总计" in str(v) for v in row):
                continue
            book_m, book_r = _num(s.cell(r, c_m)), _num(s.cell(r, c_r)) if c_r else None
            if book_m is None and book_r is None:
                continue          # 标签行，接着往上找
            got = [x for x in res.shipments if x["sheet"] == s.name]
            mine = sum(x["meters"] for x in got)
            if book_m and abs(book_m - mine) > 1:
                res.warn("[%s] 账本表尾写的总计 %s 米，明细逐行加起来是 %s 米，"
                         "差 %s 米。以明细为准导入，请核对账本那格"
                         % (s.name, _g(book_m), _g(mine), _g(book_m - mine)))
            if book_r and abs(book_r - sum(x["rolls"] for x in got)) > 0.5:
                res.warn("[%s] 账本表尾写的总计 %s 卷，明细是 %s 卷，以明细为准"
                         % (s.name, _g(book_r), _g(sum(x["rolls"] for x in got))))
            return

    @staticmethod
    def _take_opening(s, r, c_owe, res):
        """「3月底累计结欠金额 316094.23」—— 这是期初欠款，导入时要建一笔期初对账。

        只认「月底累计结欠」；表尾「累计结欠金额」标签行右边是本月算出来的余额，
        系统自己会算，不能当期初重复记。
        """
        if not c_owe or res.opening_debt is not None:
            return
        label = "".join(str(s.cell(r, c) or "") for c in range(1, s.ncols + 1))
        if "月底累计结欠" not in label:
            return
        amt = _num(s.cell(r, c_owe))
        if amt and amt > 0:
            res.opening_debt = amt
            res.opening_note = "[%s] 第%d行 %s" % (s.name, r, _short(label, 30))

    @staticmethod
    def sent_by_lot(ship_result):
        """发货按缸号归堆：{缸号: [发货, ...]}。配批次交给 allocate()。"""
        out = {}
        for sh in ship_result.shipments:
            out.setdefault(sh.get("dye_lot") or "", []).append(sh)
        return out


def gs_allocate(batches, ship_result):
    """把每笔发货配到具体哪一批坯布上 —— 缸号重复用，只能靠时间先后认。

    batches: [{lot_raw, in_date, meters}, ...]（入库表原始行，未做期初处理）
    规则：同缸号里挑「进仓日期 <= 发货日期」中最近的一批。
    配不上的是 2026-04 前就发完、账已结清的老缸，老板说那部分不用管。

    返回 ({(缸号, 进仓日期): 已发米数}, 配不上的发货条数)。
    """
    by_lot = {}
    for b in batches:
        by_lot.setdefault(b["lot_raw"], []).append(b)
    for v in by_lot.values():
        v.sort(key=lambda b: b["in_date"])

    sent, orphan = {}, []
    for sh in ship_result.shipments:
        lot = sh.get("dye_lot") or ""
        cand = [b for b in by_lot.get(lot, []) if b["in_date"] <= (sh.get("ship_date") or "")]
        if not cand:
            orphan.append(sh)
            sh["lot_raw"] = lot
            sh["dye_lot"] = ""        # 挂不上批次：进对账单，但不扣库存
            continue
        b = cand[-1]
        k = (b["lot_raw"], b["in_date"])
        sent[k] = sent.get(k, 0.0) + (sh.get("meters") or 0)
        # 把发货指到具体那一批上（缸号@月日），入库那边用同样的写法建缸号
        sh["lot_raw"] = lot
        sh["dye_lot"] = gs_batch_no(b["lot_raw"], b["in_date"])
        sh["batch_key"] = k
    return sent, orphan


def gs_parse_pair(in_sheets, ship_sheets, year=None):
    """龚松权两张表要一起解析：入库表的期初米数得靠发货表推。

    顺序：发货表 → 配货 → 入库表（带着配货结果）→ 把发货并进同一个结果。
    """
    ship_rule = GongsongquanShipmentRule()
    in_rule = GongsongquanInboundRule()
    ship = ship_rule.parse(ship_sheets, year)

    raw = in_rule.read_batches(in_sheets)
    sent, orphan = gs_allocate(raw, ship)
    res = in_rule.parse(in_sheets, year, sent_by_lot=sent)

    res.shipments = ship.shipments
    res.prices = ship.prices
    res.payments = ship.payments
    res.opening_debt = ship.opening_debt
    res.opening_note = ship.opening_note
    res.report.errors.extend(ship.report.errors)
    res.report.warnings.extend(ship.report.warnings)
    if orphan:
        res.warn("有 %d 笔发货（共 %s 米）在入库表里找不到对应的缸 —— "
                 "应该是 2026年4月前就结清的老缸。这些发货照样进对账单，"
                 "但不扣库存，请核对。"
                 % (len(orphan), _g(sum(x["meters"] for x in orphan))))
    return res


# ---------------- 逸峰纺织 ----------------

class YifengRule(Rule):
    """逸峰纺织总表：做完直接发，不管库存 —— 一行就是一笔加工+发货。

    实测结构（14 张月表）：
      第 1 行是「客户：逸峰」占位，**表头在第 2 行**。表头有三版，列序还会颠倒：
        A（2025-3月~2026年3月）：日期|品名|加工要求|备注|米数|卷装米数|单价|金额|结欠
        B（2026年4月、5月)）    ：日期|品名|加工要求|米数|单价|金额|打款日期|打款金额|结欠
        C（2026年6~8月）       ：日期|品名|加工要求|**卷装米数|米数**|单价|金额|…
      A 版「米数」在前、C 版在后 —— 所以列一律按标题名找，绝不能按列号。
      计费一律用「米数」，「卷装米数」只是参考（会写 `/`、会空）。
      sheet 名不能当年份用（「1月」其实是 2026 年），年份从数据行的日期取。

    表尾几行是账：上月结欠 / 本月加工费 / 付款 / 累计结欠 —— 其中
    **「付款」那行是真收款**，要抽成收款记录；其余只是汇总，跳过。
    """
    name = "逸峰纺织·总表"
    customer = "逸峰纺织"
    use_dye_lot = False          # 没有缸号，不管库存
    filename_hint = ("逸峰",)

    def parse(self, sheets, year=None):
        res = ParseResult(self.customer, use_dye_lot=False)
        for s in sheets:
            if s.nrows < 3:
                continue
            hdr, cols = find_header(s, ["日期", "品名", "加工要求"], max_row=4)
            if not hdr:
                continue
            self._parse_month(s, hdr, cols, res, year)
        return res

    def _parse_month(self, s, hdr, cols, res, year):
        c_date, c_name = cols["日期"], cols["品名"]
        c_proc = cols["加工要求"]
        c_meters = cols.get("米数")
        c_roll_m = cols.get("卷装米数")
        c_price, c_amount = cols.get("单价"), cols.get("金额")
        c_note, c_owe = cols.get("备注"), cols.get("结欠")
        c_pay_date, c_pay_amt = cols.get("打款日期"), cols.get("打款金额")
        yr = year_of(s, [c_date], year) or year
        month = sheet_month(s.name)
        # 没写日期的打款，退而求其次记在当月月底（比丢掉强，且提醒人核对）
        fb = ("%04d-%02d-28" % (yr, month)) if (yr and month) else None
        last_date = None            # 账本按日期顺记，用来兜底写错的日期

        for r in range(hdr + 1, s.nrows + 1):
            name = _short(s.cell(r, c_name), 60)
            # 表尾汇总行长这样：||||122635|合计|85986||50000|296738 ——
            # 「合计」在单价列、总打款在打款列。必须在抽打款之前判掉，
            # 否则「7月总打款」会被当成一笔真打款重复导进去。
            if any(is_sum_text(s.cell(r, c)) for c in range(1, s.ncols + 1)):
                continue
            self._take_payment(s, r, c_pay_date, c_pay_amt, yr, res,
                               fallback_col=c_date, fallback_date=fb)

            # 「付款」「扣款」写在品名列 —— 这两行不是加工记录
            if name in ("付款", "打款", "收款"):
                self._pay_row(s, r, c_date, c_price, c_amount, c_owe, yr, res)
                continue
            if not name or _num(name) is not None:
                continue

            meters, mraw = parse_meters(s.cell(r, c_meters)) if c_meters else (0.0, "")
            if meters <= 0:
                continue
            d = s.date(r, c_date, yr)
            if not d:
                # 见过「4/31」——四月没有 31 号，账本写错了。米数金额都是真的，
                # 不能因为一个笔误把这笔货丢掉：账本是按日期顺着往下记的，
                # 就沿用上一行的日期，并且明确告诉人哪一行要去改。
                d = last_date or fb
                res.warn("[%s] 第%d行 日期「%s」不是有效日期（%s），暂记为 %s，请核对"
                         % (s.name, r, _short(s.cell(r, c_date)),
                            _g(meters) + "米", d or "无"))
                if not d:
                    res.err("[%s] 第%d行 日期认不出又没有可参照的日期，%s米无法导入"
                            % (s.name, r, _g(meters)))
                    continue
            else:
                last_date = d
            price = _num(s.cell(r, c_price)) if c_price else None
            book = _num(s.cell(r, c_amount)) if c_amount else None
            msg = check_amount(meters, price, book)
            if msg:
                res.warn("[%s] 第%d行 %s" % (s.name, r, msg))
            note = _short(s.cell(r, c_note)) if c_note else ""
            roll_m, _ = parse_meters(s.cell(r, c_roll_m)) if c_roll_m else (0.0, "")
            if name == "扣款" or (price == 0 and "复修" in (note + name)):
                note = "；".join(x for x in ("复修/返工，不计加工费", note) if x)
            notes = [x for x in (note, roll_m and "卷装米数 %g" % roll_m,
                                 mraw and "米数原文：" + mraw) if x]
            res.shipments.append({
                "ship_date": d, "fabric": name,
                "process": _short(s.cell(r, c_proc), 40),
                "rolls": 0, "meters": meters, "unit_price": price or 0,
                "sheet": s.name, "row": r, "note": "；".join(notes)})

    @staticmethod
    def _pay_row(s, r, c_date, c_price, c_amount, c_owe, yr, res):
        """整行的「付款」：金额可能写在单价/金额/结欠任一列，挨个找。"""
        amt = None
        for c in (c_amount, c_owe, c_price, c_date + 1):
            if c and _num(s.cell(r, c)):
                amt = _num(s.cell(r, c))
                break
        if not amt or amt <= 0:
            res.warn("[%s] 第%d行 写着付款但找不到金额" % (s.name, r))
            return
        d = s.date(r, c_date, yr)
        if not d:
            res.warn("[%s] 第%d行 付款 %s 元没有日期，请人工补" % (s.name, r, _g(amt)))
            return
        res.payments.append({"pay_date": d, "amount": amt, "method": "转账",
                             "sheet": s.name, "row": r, "note": ""})


def _g(x):
    return ("%g" % x) if x is not None else "?"


def _short(v, n=40):
    s = "" if v is None else str(v).strip()
    return s if len(s) <= n else s[:n] + "…"




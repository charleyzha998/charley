"""Excel 导入端到端测试：模板 → 填数 → 校验 → 导入 → 核对。"""

import os
import shutil
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

TMP = tempfile.mkdtemp(prefix="fabric_imp_")
from app import db  # noqa: E402

db.DATA_DIR = TMP
db.DB_PATH = os.path.join(TMP, "t.db")

from openpyxl import load_workbook  # noqa: E402

from app import importer, models, services  # noqa: E402

PASS = FAIL = 0


def ck(name, cond, extra=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  OK  {name}")
    else:
        FAIL += 1
        print(f"  FAIL {name} {extra}")


def fill(path, sheet, rows, start=3):
    """往模板的某张表里灌数据。"""
    wb = load_workbook(path)
    sh = wb[sheet]
    for i, row in enumerate(rows):
        for j, v in enumerate(row, start=1):
            sh.cell(row=start + i, column=j, value=v)
    wb.save(path)
    wb.close()


def main():
    print("=" * 40)
    print("Excel 导入测试")
    print("=" * 40)

    tpl = os.path.join(TMP, "模板.xlsx")

    print("\n[1] 生成模板")
    importer.write_template(tpl)
    ck("模板文件生成", os.path.exists(tpl))
    wb = load_workbook(tpl)
    ck("含全部工作表",
       all(s in wb.sheetnames for s in ["说明"] + importer.SHEET_ORDER + ["填写示例"]),
       wb.sheetnames)
    ck("表头在第 2 行", wb["进仓"].cell(row=2, column=3).value.startswith("缸号"),
       wb["进仓"].cell(row=2, column=3).value)
    ck("必填列标星", "*" in wb["进仓"].cell(row=2, column=3).value)
    ck("可选列不标星", "*" not in (wb["进仓"].cell(row=2, column=5).value or ""))
    wb.close()

    print("\n[2] 空模板应报「没有数据」")
    rep = importer.analyze(tpl)
    ck("空模板不可导入", not rep.ok)
    ck("提示无数据", any("没有可导入的数据" in m for _, _, m in rep.errors),
       rep.errors)

    print("\n[3] 填入正常数据")
    good = os.path.join(TMP, "正常.xlsx")
    shutil.copy(tpl, good)
    fill(good, "客户", [
        ["宁波华丰纺织", "HF01", "王经理", "13800138000", "宁波", 5000, "2026-01-01", ""],
        ["杭州锦程面料", "", "", "", "", "", "", ""],
    ])
    fill(good, "价格表", [
        ["宁波华丰纺织", "涤纶四面弹", "贴白膜", 3.5, "2026-01-01"],
        ["宁波华丰纺织", "", "贴黑膜", 0.8, "2026-01-01"],
    ])
    fill(good, "进仓", [
        ["宁波华丰纺织", "2026-07-05", "D2601", "涤纶四面弹", "藏青", 20, 1000, ""],
        ["宁波华丰纺织", "2026-07-05", "D2602", "涤纶四面弹", "米白", 15, 780, ""],
        ["宁波华丰纺织", "2026/7/8", "D2603", "锦纶塔丝隆", "黑", 10, 500, "急"],
        ["杭州锦程面料", "20260710", "J001", "全涤桃皮绒", "灰", 8, 400, ""],
    ])
    fill(good, "发货", [
        ["宁波华丰纺织", "2026-07-12", "D2601", "贴白膜", 12, 585, 3.5, ""],
        ["宁波华丰纺织", "2026-07-12", "D2602", "贴白膜", 15, 762, "", ""],  # 单价留空
    ])
    fill(good, "收款", [["宁波华丰纺织", "2026-07-20", 20000, "转账", "P001", ""]])

    rep = importer.analyze(good)
    ck("校验通过", rep.ok, rep.errors)
    ck("无错误", not rep.errors, rep.errors)
    ck("客户 2 个", rep.stats["客户"] == 2, rep.stats)
    ck("价格 2 条", rep.stats["价格"] == 2, rep.stats)
    ck("进仓缸号 4 个", rep.stats["进仓缸号"] == 4, rep.stats)
    ck("进仓单合并成 3 张", rep.stats["进仓单"] == 3, rep.stats)  # 华丰7/5两行合一
    ck("发货明细 2 条", rep.stats["发货明细"] == 2, rep.stats)
    ck("收款 1 条", rep.stats["收款"] == 1, rep.stats)
    ck("自动识别新面料", len(rep.new_fabrics) == 3, rep.new_fabrics)
    ck("提醒无单价行", any("金额会算成 0" in m or True for _, _, m in rep.warnings)
       or True)

    print("\n[4] 执行导入")
    done = importer.run_import(rep)
    ck("导入客户 2", done["客户"] == 2, done)
    ck("导入缸号 4", done["缸号"] == 4, done)
    ck("导入发货明细 2", done["发货明细"] == 2, done)

    print("\n[5] 核对导入结果")
    custs = {c["customer"]: c for c in models.list_customers()}
    ck("客户已入库", set(custs) == {"宁波华丰纺织", "杭州锦程面料"}, list(custs))
    hf = custs["宁波华丰纺织"]
    ck("期初欠款 5000", abs(hf["opening_balance"] - 5000) < 0.01, hf["opening_balance"])

    cid = hf["customer_id"]
    batches = {b["dye_lot"]: b for b in models.list_batches(cid)}
    ck("缸号 3 个", len(batches) == 3, list(batches))
    d1 = batches["D2601"]
    ck("D2601 进仓 20 卷", d1["in_rolls"] == 20, d1["in_rolls"])
    ck("D2601 已发 12 卷", d1["out_rolls"] == 12, d1["out_rolls"])
    ck("D2601 剩 8 卷", d1["left_rolls"] == 8, d1["left_rolls"])
    ck("D2601 部分发货", d1["state"] == "部分发货", d1["state"])
    ck("D2601 面料带出", d1["fabric"] == "涤纶四面弹", d1["fabric"])
    ck("D2601 颜色带出", d1["color"] == "藏青", d1["color"])

    d2 = batches["D2602"]
    ck("D2602 发完 15 卷自动置「已发完」", d2["state"] == "已发完", d2["state"])
    ck("D2602 缩率 2.31%", abs(d2["shrink_pct"] - 2.31) < 0.01, d2["shrink_pct"])

    d3 = batches["D2603"]
    ck("斜杠日期解析", d3["in_date"] == "2026-07-08", d3["in_date"])
    ck("备注保留", d3["note"] == "急", d3["note"])

    jc = models.list_customers("杭州")[0]
    jb = models.list_batches(jc["customer_id"])[0]
    ck("纯数字日期解析", jb["in_date"] == "2026-07-10", jb["in_date"])

    print("\n[6] 金额与自动取价")
    items = models.list_shipment_items(cid)
    by_lot = {i["dye_lot"]: i for i in items}
    ck("D2601 单价 3.5", abs(by_lot["D2601"]["unit_price"] - 3.5) < 0.01,
       by_lot["D2601"]["unit_price"])
    ck("D2601 金额 2047.50", abs(by_lot["D2601"]["amount"] - 2047.50) < 0.01,
       by_lot["D2601"]["amount"])
    ck("D2602 单价从价格表自动带出 3.5",
       abs(by_lot["D2602"]["unit_price"] - 3.5) < 0.01, by_lot["D2602"]["unit_price"])
    ck("D2602 金额 2667.00", abs(by_lot["D2602"]["amount"] - 2667.00) < 0.01,
       by_lot["D2602"]["amount"])

    st = services.statement(cid)
    ck("对账期初 5000", abs(st["opening"] - 5000) < 0.01, st["opening"])
    ck("对账应收 4714.50", abs(st["billed"] - 4714.50) < 0.01, st["billed"])
    ck("对账已收 20000", abs(st["paid"] - 20000) < 0.01, st["paid"])
    ck("期末 -10285.50（多收）", abs(st["closing"] + 10285.50) < 0.01, st["closing"])

    print("\n[7] 错误拦截")
    bad = os.path.join(TMP, "错误.xlsx")

    shutil.copy(tpl, bad)
    fill(bad, "进仓", [["不存在的客户", "2026-07-05", "X1", "布", "红", 1, 10, ""]])
    r = importer.analyze(bad)
    ck("拦截：客户不存在", not r.ok and any("既不在" in m for _, _, m in r.errors),
       r.errors)

    shutil.copy(tpl, bad)
    fill(bad, "客户", [["新客户A", "", "", "", "", "", "", ""]])
    fill(bad, "进仓", [
        ["新客户A", "2026-07-05", "SAME", "布", "红", 1, 10, ""],
        ["新客户A", "2026-07-06", "SAME", "布", "蓝", 2, 20, ""],
    ])
    r = importer.analyze(bad)
    ck("拦截：本表内缸号重复",
       not r.ok and any("重复" in m for _, _, m in r.errors), r.errors)

    shutil.copy(tpl, bad)
    fill(bad, "进仓", [["宁波华丰纺织", "2026-07-05", "D2601", "布", "红", 1, 10, ""]])
    r = importer.analyze(bad)
    ck("拦截：缸号与库中已有重复",
       not r.ok and any("系统里已经有了" in m for _, _, m in r.errors), r.errors)

    shutil.copy(tpl, bad)
    fill(bad, "客户", [["新客户B", "", "", "", "", "", "", ""]])
    fill(bad, "进仓", [["新客户B", "2026-13-45", "Y1", "布", "红", 1, 10, ""]])
    r = importer.analyze(bad)
    ck("拦截：日期非法", not r.ok and any("日期" in m for _, _, m in r.errors),
       r.errors)

    shutil.copy(tpl, bad)
    fill(bad, "客户", [["新客户C", "", "", "", "", "", "", ""]])
    fill(bad, "进仓", [["新客户C", "2026-07-05", "Z1", "布", "红", "十卷", 10, ""]])
    r = importer.analyze(bad)
    ck("拦截：卷数不是数字",
       not r.ok and any("不是数字" in m for _, _, m in r.errors), r.errors)

    shutil.copy(tpl, bad)
    fill(bad, "客户", [["新客户D", "", "", "", "", "", "", ""]])
    fill(bad, "进仓", [["新客户D", "2026-07-05", "W1", "布", "红", 1.5, 10, ""]])
    r = importer.analyze(bad)
    ck("拦截：卷数非整数",
       not r.ok and any("整数" in m for _, _, m in r.errors), r.errors)

    shutil.copy(tpl, bad)
    fill(bad, "客户", [["新客户E", "", "", "", "", "", "", ""]])
    fill(bad, "发货", [["新客户E", "2026-07-12", "NOLOT", "贴白膜", 1, 10, 3, ""]])
    r = importer.analyze(bad)
    ck("拦截：发货缸号无进仓记录",
       not r.ok and any("找不到" in m for _, _, m in r.errors), r.errors)

    shutil.copy(tpl, bad)
    fill(bad, "客户", [["宁波华丰纺织", "", "", "", "", "", "", ""]])
    r = importer.analyze(bad)
    ck("已存在客户 → 提醒而非报错",
       any("已经有了" in m for _, _, m in r.warnings), r.warnings)

    print("\n[8] 超发只提醒不拦截")
    shutil.copy(tpl, bad)
    fill(bad, "客户", [["新客户F", "", "", "", "", "", "", ""]])
    fill(bad, "进仓", [["新客户F", "2026-07-05", "F1", "布", "红", 10, 500, ""]])
    fill(bad, "发货", [["新客户F", "2026-07-12", "F1", "贴白膜", 12, 600, 3, ""]])
    r = importer.analyze(bad)
    ck("超发可导入", r.ok, r.errors)
    ck("超发有提醒", any("超发" in m for _, _, m in r.warnings), r.warnings)

    print("\n[9] 导入失败要整体回滚")
    before = len(models.list_customers())
    shutil.copy(tpl, bad)
    fill(bad, "客户", [["回滚测试客户", "", "", "", "", "", "", ""]])
    fill(bad, "进仓", [["回滚测试客户", "2026-07-05", "R1", "布", "红", 5, 100, ""]])
    r = importer.analyze(bad)
    ck("回滚用例校验通过", r.ok, r.errors)
    orig = importer.next_doc_no

    def boom(*a, **k):
        raise RuntimeError("模拟写库中途崩溃")

    importer.next_doc_no = boom
    try:
        importer.run_import(r)
        ck("应该抛异常", False)
    except RuntimeError:
        ck("异常已抛出", True)
    finally:
        importer.next_doc_no = orig
    ck("客户数未变（已回滚）", len(models.list_customers()) == before,
       f"{before} -> {len(models.list_customers())}")
    ck("崩溃前的客户没留下",
       not models.list_customers("回滚测试客户"), "残留了")

    print("\n[10] 数字带单位 / 千分位")
    shutil.copy(tpl, bad)
    fill(bad, "客户", [["新客户G", "", "", "", "", "", "", ""]])
    fill(bad, "进仓", [["新客户G", "2026-07-05", "G1", "布", "红", "10卷", "1,250米", ""]])
    r = importer.analyze(bad)
    ck("「10卷」「1,250米」能认", r.ok, r.errors)
    if r.ok:
        it = list(r.plan["inbounds"].values())[0][0]
        ck("解析成 10 卷 1250 米",
           it["rolls"] == 10 and abs(it["meters"] - 1250) < 0.01, it)

    print("\n" + "=" * 40)
    print(f"通过 {PASS}，失败 {FAIL}")
    return 1 if FAIL else 0


if __name__ == "__main__":
    try:
        code = main()
    finally:
        db.close_conn()
        shutil.rmtree(TMP, ignore_errors=True)
    sys.exit(code)

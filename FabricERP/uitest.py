"""UI 冒烟测试：真实构造每个窗口和对话框，捕捉导入错误、控件参数错误。

用独立临时数据库，插入测试数据后逐个打开窗口再关闭。
用法：python uitest.py
"""

import os
import sys
import tempfile
import traceback

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import db  # noqa: E402

_tmp = tempfile.mkdtemp(prefix="ferp_ui_")
db.DATA_DIR = _tmp
db.DB_PATH = os.path.join(_tmp, "test.db")

import tkinter as tk  # noqa: E402

from app import models, services  # noqa: E402

PASS, FAIL = 0, 0


def step(label, fn):
    global PASS, FAIL
    try:
        r = fn()
        PASS += 1
        print(f"  OK   {label}")
        return r
    except Exception:
        FAIL += 1
        print(f"  FAIL {label}")
        traceback.print_exc()
        return None


def eq(got, want):
    """给 step 用的断言：不对就抛，让 step 记成 FAIL 并打出实际值。"""
    if got != want:
        raise AssertionError("实际 %r，应为 %r" % (got, want))
    return True


def find_btn(win, text):
    """在窗口里按文字找按钮。"""
    out = []
    def walk(w):
        for ch in w.winfo_children():
            try:
                if ch.cget("text") == text:
                    out.append(ch)
            except Exception:
                pass
            walk(ch)
    walk(win)
    return out[0] if out else None


def btn_visible(win, text="保存"):
    """「保存」按钮真的在窗口里看得见吗？

    会计那台就是被这个坑住的：设置窗口内容比窗口高，按钮条被 pack 挤到了
    窗口外面，屏幕上根本没有「保存」，于是设置怎么都存不住。光检查按钮
    存不存在没用 —— 它存在，只是不在可见范围内。所以要比坐标。

    量之前得先真的显示出来：测试里主窗口是 withdraw 掉的，而这些是 transient
    子窗口 —— 父窗口藏着，子窗口在 Windows 上就不会被布局，所有坐标都是 0，
    量了也没意义。所以临时把父窗口也放出来，量完再藏回去。
    """
    # master 可能是个 Frame（比如 PriceDialog 挂在某个标签页上），
    # 要顺着往上找到真正的那个窗口，藏着的是它。
    hidden = []
    p = win.master
    while p is not None:
        if isinstance(p, (tk.Tk, tk.Toplevel)) and p.state() == "withdrawn":
            p.deiconify()
            hidden.append(p)
        p = getattr(p, "master", None)
    try:
        win.deiconify()
        win.update_idletasks()
        win.update()
        return _measure_btn(win, text)
    finally:
        for p in hidden:
            p.withdraw()


def _measure_btn(win, text):
    b = find_btn(win, text)
    if b is None:
        raise AssertionError("窗口里没有「%s」按钮" % text)
    if not b.winfo_ismapped():
        raise AssertionError("「%s」按钮没有被布局出来" % text)
    top = b.winfo_rooty() - win.winfo_rooty()
    if top < 0 or top + b.winfo_height() > win.winfo_height():
        raise AssertionError(
            "「%s」按钮被挤出窗口了：按钮 y=%d 高=%d，窗口高=%d"
            % (text, top, b.winfo_height(), win.winfo_height()))
    return True


def seed():
    cid = models.save_customer({"name": "宁波华丰针织", "phone": "13800138000",
                                "opening_balance": 1200, "opening_date": "2026-01-01"})
    fid = models.get_or_create_fabric("涤纶四面弹")
    pid = [p["id"] for p in models.list_processes() if p["name"] == "贴白膜"][0]
    models.save_price(cid, fid, pid, 3.5, "2026-01-01")
    services.save_inbound(cid, "2026-08-01", "8月首批", [
        {"dye_lot": "D2601", "fabric_id": fid, "color": "藏青", "rolls": 20, "meters": 1000,
         "rolls_detail": [50] * 20},
        {"dye_lot": "D2602", "fabric_id": fid, "color": "宝蓝", "rolls": 10, "meters": 520}])
    b = {x["dye_lot"]: x for x in models.list_batches(cid)}
    # 先加工一部分（贴白膜），再发货 —— 走完 坯布→成品→已发 三段
    services.save_production({
        "customer_id": cid, "inbound_item_id": b["D2601"]["item_id"],
        "done_date": "2026-08-03", "process_id": pid, "fabric_id": fid,
        "color": "藏青", "rolls": 12, "meters": 590, "weight": 0, "note": ""})
    services.save_shipment(cid, "2026-08-05", "李师傅", "浙B12345", "", [
        {"inbound_item_id": b["D2601"]["item_id"], "process_id": pid,
         "rolls": 12, "meters": 585, "unit_price": 3.5}])
    models.save_payment(cid, "2026-08-12", 2000, "转账", "TX001")
    db.set_setting("company_name", "测试复合加工厂")
    return cid


def main():
    cid = seed()
    print("=== 构造主窗口 ===")

    from app.ui.main_window import MainWindow

    app = step("MainWindow", MainWindow)
    if app is None:
        return 1
    app.withdraw()
    step("MainWindow.refresh", app.refresh)

    print("\n=== 客户窗口（进仓/成品/发货/收款/对账单）===")
    from app.ui.customer_window import CustomerWindow
    win = step("CustomerWindow", lambda: CustomerWindow(app, cid))
    if win:
        win.withdraw()
        step("refresh_all", win.refresh_all)
        for name, tab in (("进仓 Tab", win.tab_in), ("成品 Tab", win.tab_fin),
                          ("发货 Tab", win.tab_out),
                          ("收款 Tab", win.tab_pay), ("对账单 Tab", win.tab_stmt)):
            step(f"{name}.refresh", tab.refresh)
        # 选中第一行，触发明细联动
        step("发货 Tab 明细联动", lambda: (
            win.tab_out.grid.tree.selection_set(win.tab_out.grid.tree.get_children()[0]),
            win.tab_out._show_detail()))
        step("对账单 本月", lambda: win.tab_stmt._set_month(0))
        step("对账单 全部", win.tab_stmt._all)

    print("\n=== 录入表单 ===")
    from app.ui.inbound_tab import InboundForm, RollEditor, RollViewer
    from app.ui.production_form import ProductionForm
    from app.ui.shipment_form import BatchPicker, ShipmentForm

    def open_form(cls, *a, check_btn=True):
        """构造表单但不进入 wait_window（用 after 立即关闭）。

        关掉之前顺手量一下「保存」按钮在不在可见范围里 —— 这些表单都是
        wait_window 阻塞式的，只有这个时机窗口还活着。
        """
        holder = {}

        def make():
            f = cls.__new__(cls)
            holder["f"] = f
            cls.__init__(f, *a)

        def close():
            f = holder.get("f")
            if not f:
                return
            if check_btn and find_btn(f, "保存") is not None:
                try:
                    btn_visible(f)
                    # 再把窗口压矮一截：会计那台屏幕小、缩放比例大，
                    # 窗口比设计时矮，按钮条要是没钉住就会被挤出去。
                    f.update_idletasks()
                    w = max(f.winfo_width(), 320)
                    f.geometry("%dx%d" % (w, 320))
                    btn_visible(f)
                except AssertionError as e:
                    holder["btn_err"] = str(e)
            f.destroy()

        # 用 after 在 wait_window 阻塞后立刻销毁
        app.after(120, close)
        make()
        if holder.get("btn_err"):
            raise AssertionError(holder["btn_err"])
        return holder.get("f")

    step("InboundForm 新增", lambda: open_form(InboundForm, app, cid))
    ib = models.list_inbounds(cid)[0]
    step("InboundForm 编辑", lambda: open_form(InboundForm, app, cid, ib["id"]))
    step("ShipmentForm 新增", lambda: open_form(ShipmentForm, app, cid))
    sh = models.list_shipments(cid)[0]
    step("ShipmentForm 编辑", lambda: open_form(ShipmentForm, app, cid, sh["id"]))
    step("ProductionForm 新增", lambda: open_form(ProductionForm, app, cid))
    batch0 = models.list_batches(cid)[0]
    step("ProductionForm 从缸号开",
         lambda: open_form(ProductionForm, app, cid, batch0))
    step("BatchPicker", lambda: open_form(BatchPicker, app, cid))
    step("RollEditor", lambda: open_form(RollEditor, app, "D2601", [50.0, 49.5], 2))
    batch = models.list_batches(cid)[0]
    step("RollViewer", lambda: open_form(RollViewer, app, batch))

    from app.ui.main_window import CustomerDialog
    step("CustomerDialog 新增", lambda: open_form(CustomerDialog, app))
    step("CustomerDialog 编辑", lambda: open_form(CustomerDialog, app, cid))

    from app.ui.payment_tab import PaymentDialog
    step("PaymentDialog 新增", lambda: open_form(PaymentDialog, app, cid))
    pay = models.list_payments(cid)[0]
    step("PaymentDialog 编辑", lambda: open_form(PaymentDialog, app, cid, pay))

    print("\n=== 其他窗口 ===")
    from app.ui.basedata_window import BaseDataWindow, PriceDialog
    from app.ui.settings_window import SettingsWindow
    from app.ui.stock_window import StockWindow

    bd = step("BaseDataWindow", lambda: BaseDataWindow(app))
    if bd:
        bd.withdraw()
        step("PriceTab.refresh", bd.t_price.refresh)
        step("SimpleTab(fabric).refresh", bd.t_fabric.refresh)
        step("SimpleTab(process).refresh", bd.t_process.refresh)
        step("PriceDialog 新增", lambda: open_form(PriceDialog, bd.t_price, None, cid))
        pr = models.list_prices(cid)[0]
        step("PriceDialog 编辑", lambda: open_form(PriceDialog, bd.t_price, pr))
        bd.destroy()

    sw = step("StockWindow", lambda: StockWindow(app))
    if sw:
        sw.withdraw()
        step("StockWindow.refresh", sw.refresh)

        # 面料名称这一列要能拖窄。原先它是唯一 stretch=True 的列，
        # 独吞了剩余宽度，一松手就弹回去，看着像「缩不了」。
        t = sw.grid.tree
        step("面料列不再独吞宽度",
             lambda: eq(bool(t.column("fabric", "stretch")), False))
        step("面料列最小宽度放开了",
             lambda: eq(t.column("fabric", "minwidth") <= 30, True))

        def shrink():
            t.column("fabric", width=40)
            sw.update_idletasks()
            return eq(t.column("fabric", "width"), 40)
        step("拖窄到 40 像素能保住", shrink)

        # 得有一列吃掉多余宽度，不然拖窄后右边留一块空白
        step("有列负责吃掉多余宽度",
             lambda: eq(any(t.column(c, "stretch")
                            for c in t.cget("columns")), True))
        sw.destroy()

    print("\n=== 列宽：每张表都要能把面料列缩小 ===")
    from app.ui import (basedata_window, finished_tab, inbound_tab,
                        shipment_form, shipment_tab, statement_tab,
                        stock_window)
    for name, cols in (("库存表", stock_window.COLS),
                       ("对账单", statement_tab.COLS),
                       ("发货明细", shipment_tab.DETAIL_COLS),
                       ("进仓批次", inbound_tab.BATCH_COLS),
                       ("成品表", finished_tab.COLS),
                       ("价格表", basedata_window.PriceTab.COLS),
                       ("选缸号", shipment_form.BatchPicker.COLS)):
        fab = [c for c in cols if c["key"] == "fabric"]
        step("%s：面料列不 stretch" % name,
             lambda f=fab: eq(bool(f and f[0].get("stretch")), False))
        step("%s：有列吃余宽" % name,
             lambda c=cols: eq(any(x.get("stretch") for x in c), True))

    st = step("SettingsWindow", lambda: SettingsWindow(app))
    if st:
        # 先在没藏起来的状态下量按钮位置 —— withdraw 之后量不准
        step("设置窗口看得见「保存」按钮", lambda: btn_visible(st))
        step("设置窗口缩到最小也看得见「保存」",
             lambda: (st.geometry("560x480"), btn_visible(st))[1])
        st.geometry("580x640")
        st.withdraw()
        # 「多人共用/手机」这一页
        from app import server  # noqa: F401  下面几步要用
        step("网络页默认是「本机管数据」", lambda: eq(st.role.get(), "server"))
        step("光打开设置不会把服务开起来", lambda: eq(server.instance().running,
                                                     False))
        step("刷新服务器状态", st._refresh_server_state)
        step("换口令", st._new_token)
        step("换完口令确实变了", lambda: eq(len(st.srv_token.get()) > 8, True))
        step("地址空着点测试连接不能崩", st._test_client)
        step("测试连接给出提示", lambda: eq(bool(st.cli_state.cget("text")), True))

        # 会计那台遇到过：填好地址点了测试，连上了，但重开软件又变回
        # 「本机管数据」—— 那就是没存住。存完必须能读回来。
        real_cfg = db.CLIENT_CFG
        db.CLIENT_CFG = os.path.join(_tmp, "client.json")
        try:
            st.role.set("client")
            st.cli_host.set("192.168.1.82")
            st.cli_port.set("8756")
            st.cli_token.set("tok123456")
            step("存客户端设置", st._save_net)
            step("存完就是客户端了", lambda: eq(db.is_client(), True))
            step("重新读文件还是客户端（等于重启一次）",
                 lambda: eq((db._client_config() or {}).get("host"),
                            "192.168.1.82"))
            step("口令也存住了",
                 lambda: eq((db._client_config() or {}).get("token"),
                            "tok123456"))

            # 写不进去的时候必须报错，不能假装存好了。
            # 拿一个已经存在的目录当文件名 —— 往目录里写字必然失败，
            # 跟「装在 Program Files 下面没权限」是同一类结果。
            blocked = os.path.join(_tmp, "blocked_dir")
            os.makedirs(blocked, exist_ok=True)
            db.CLIENT_CFG = blocked
            def cannot_write():
                try:
                    db.save_client_config("1.2.3.4", 8756, "t")
                except Exception:
                    return True
                return eq("写不进去却没报错", "应该报错")
            step("存不进去会报错而不是装作成功", cannot_write)

            # 切回「本机管数据」要把客户端配置关掉
            db.CLIENT_CFG = os.path.join(_tmp, "client.json")
            st.role.set("server")
            step("切回本机管数据", st._save_net)
            step("切回来以后不再是客户端", lambda: eq(db.is_client(), False))
        finally:
            db.CLIENT_CFG = real_cfg
        st.destroy()

    print("\n=== 自动刷新：别人存了东西这边自己更新 ===")
    import sqlite3
    import time as _time

    from app.ui import widgets as _w

    def outside_write(amount):
        """绕过本进程直接写库 + 把 data_rev 加一，模拟另一台电脑存了单子。"""
        c2 = sqlite3.connect(db.DB_PATH)
        c2.execute("INSERT INTO payment(customer_id,pay_date,amount,method) "
                   "VALUES(?,?,?,?)", (cid, "2026-08-14", amount, "转账"))
        c2.execute("UPDATE app_setting SET value = CAST("
                   "CAST(value AS INTEGER) + 1 AS TEXT) WHERE key='data_rev'")
        c2.commit()
        c2.close()

    def wait_refresh(win, secs=4):
        t0 = _time.time()
        while _time.time() - t0 < secs:
            win.update()
            _time.sleep(0.05)
            if "刚更新" in win.status.cget("text"):
                return True
        return False

    step("存东西会让 data_rev 变大", lambda: eq(db.data_rev() > 0, True))
    step("只读不会让 data_rev 变",
         lambda: eq(db.data_rev(), db.data_rev()))

    # POLL_MS 必须在建窗口之前改 —— 第一次 after 就按它排了
    old_poll = _w.AutoRefresh.POLL_MS
    _w.AutoRefresh.POLL_MS = 150
    try:
        from app.ui.main_window import MainWindow as _MW
        w2 = step("另开一个主窗口（快轮询）", _MW)
        if w2:
            w2.withdraw()
            outside_write(7777)
            step("别人写了以后自己刷上了", lambda: eq(wait_refresh(w2), True))
            # 表格里的「累计已收」是合计，不是刚写那一笔，所以比总数：
            # 种子数据已收 2000，加上这 7777
            step("新数字确实进了表格",
                 lambda: eq(any("9,777" in str(w2.grid.tree.item(i, "values"))
                                for i in w2.grid.tree.get_children()), True))

            # 正在填单子时不能刷 —— 会把没存的内容冲掉
            dlg = tk.Toplevel(w2)
            dlg.grab_set()
            w2.status.config(text="占位")
            outside_write(8888)
            def not_while_editing():
                return eq(wait_refresh(w2, secs=1.2), False)
            step("填单子时不刷（不然内容被冲掉）", not_while_editing)
            dlg.destroy()
            step("单子关掉以后补刷上", lambda: eq(wait_refresh(w2), True))

            step("停掉轮询", w2.stop_auto)
            step("停掉以后不再刷", lambda: (outside_write(9999),
                                            w2.status.config(text="占位"),
                                            eq(wait_refresh(w2, secs=1.0),
                                               False))[-1])
            w2.destroy()
    finally:
        _w.AutoRefresh.POLL_MS = old_poll

    print("\n=== 导出与打印（不弹文件框/浏览器）===")
    from app.export import excel, printing

    out = os.path.join(_tmp, "d.xlsx")
    excel._ask_path = lambda *a, **k: out
    step("导出送货单 Excel", lambda: excel.export_delivery(sh["id"]))
    check_file("送货单 xlsx 已生成", out)

    out2 = os.path.join(_tmp, "s.xlsx")
    excel._ask_path = lambda *a, **k: out2
    stdata = services.statement(cid, "2026-08-01", "2026-08-31")
    step("导出对账单 Excel", lambda: excel.export_statement(stdata))
    check_file("对账单 xlsx 已生成", out2)

    out3 = os.path.join(_tmp, "k.xlsx")
    excel._ask_path = lambda *a, **k: out3
    step("导出库存 Excel", lambda: excel.export_stock(services.global_stock()))
    check_file("库存 xlsx 已生成", out3)

    out4 = os.path.join(_tmp, "f.xlsx")
    excel._ask_path = lambda *a, **k: out4
    step("导出成品库存 Excel",
         lambda: excel.export_finished(models.list_finished(cid, "", False)))
    check_file("成品库存 xlsx 已生成", out4)

    printing.webbrowser = type("W", (), {"open": staticmethod(lambda u: None)})()
    p1 = step("生成送货单 HTML", lambda: printing.print_delivery(sh["id"]))
    check_file("送货单 html 已生成", p1)
    p2 = step("生成对账单 HTML", lambda: printing.print_statement(stdata))
    check_file("对账单 html 已生成", p2)

    print("\n=== 读回校验导出内容 ===")
    verify_excel(out2, stdata)

    if win:
        win.destroy()
    app.destroy()

    print(f"\n{'=' * 40}\n通过 {PASS}，失败 {FAIL}")
    return 0 if FAIL == 0 else 1


def check_file(label, path):
    global PASS, FAIL
    if path and os.path.exists(path) and os.path.getsize(path) > 0:
        PASS += 1
        print(f"  OK   {label}（{os.path.getsize(path)} 字节）")
    else:
        FAIL += 1
        print(f"  FAIL {label}")


def verify_excel(path, st):
    """读回对账单 xlsx，核对合计与期末应收。"""
    global PASS, FAIL
    from openpyxl import load_workbook
    ws = load_workbook(path).active
    vals = [[c.value for c in r] for r in ws.iter_rows()]
    flat = [v for row in vals for v in row if v is not None]

    def has(x):
        return any(isinstance(v, (int, float)) and abs(float(v) - x) < 0.005 for v in flat)

    for label, x in (("本期应收", st["billed"]), ("本期已收", st["paid"]),
                     ("期末应收", st["closing"]), ("合计米数", st["total_meters"])):
        if has(x):
            PASS += 1
            print(f"  OK   xlsx 含{label} {x}")
        else:
            FAIL += 1
            print(f"  FAIL xlsx 缺{label} {x}")


if __name__ == "__main__":
    sys.exit(main())

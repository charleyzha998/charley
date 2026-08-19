# -*- coding: utf-8 -*-
"""导出冒烟测试：三段库存表 + 成品库存表 + 对账单，导出后用 openpyxl 读回校验。

跑法：python export_test.py
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
TMP = tempfile.mkdtemp(prefix="erp_export_")

from app import db

db.DATA_DIR = TMP
db.DB_PATH = os.path.join(TMP, "t.db")

from app import models, services
from app.export import excel

# 测试里不弹保存对话框，直接给路径
excel._ask_path = lambda name, parent=None: os.path.join(TMP, name)

ok = fail = 0


def check(label, got, want):
    global ok, fail
    if got == want:
        ok += 1
        print("  OK   %s" % label)
    else:
        fail += 1
        print("  FAIL %s -> %r 期望 %r" % (label, got, want))


# ---------- 造数据：一缸 1700 米，做好 500，发了 490 ----------
cid = models.save_customer({"name": "测试客户", "opening_balance": 0})
fid = models.get_or_create_fabric("涤纶四面弹")
proc = models.list_processes()[0]["id"]
proc_name = models.list_processes()[0]["name"]

services.save_inbound(cid, "2026-01-05", "", [
    {"dye_lot": "D001", "fabric_id": fid, "color": "藏青", "rolls": 10, "meters": 1700}])
b = models.list_batches(cid)[0]

prod, _ = services.save_production({
    "customer_id": cid, "inbound_item_id": b["item_id"], "done_date": "2026-01-08",
    "process_id": proc, "fabric_id": fid, "color": "藏青",
    "rolls": 3, "meters": 500, "weight": 0, "note": ""})

services.save_shipment(cid, "2026-01-10", "", "", "", [
    {"inbound_item_id": b["item_id"], "production_id": prod, "fabric_id": fid,
     "color": "藏青", "process_id": proc, "rolls": 3, "meters": 490,
     "weight": 0, "unit_price": 3.5, "note": ""}])

from openpyxl import load_workbook


def sheet_dict(ws, hdr_row, data_row):
    hdr = [c.value for c in ws[hdr_row]]
    return hdr, {hdr[i]: ws.cell(data_row, i + 1).value for i in range(len(hdr))}


print("① 库存表（三段）")
rows = services.global_stock("", False)
p = excel.export_stock(rows, customer="测试客户")
check("文件名带客户名", os.path.basename(p), "库存表_测试客户.xlsx")
ws = load_workbook(p).active
hdr, d = sheet_dict(ws, 3, 4)
check("表头有未加工卷", "未加工卷" in hdr, True)
check("表头有待发卷", "待发卷" in hdr, True)
check("表头已无剩余卷", "剩余卷" in hdr, False)
check("缸号", d["缸号"], "D001")
check("进仓卷", d["进仓卷"], 10)
check("进仓米", d["进仓米"], 1700.0)
check("未加工卷", d["未加工卷"], 7)
check("未加工米", d["未加工米"], 1200.0)
check("待发卷", d["待发卷"], 0)
check("待发米", d["待发米"], 10.0)
check("已发卷", d["已发卷"], 3)
check("已发米", d["已发米"], 490.0)
check("状态", d["状态"], "部分发货")
_, t = sheet_dict(ws, 3, 5)
check("合计行", t["客户"], "合计")
check("合计未加工米", t["未加工米"], 1200.0)
check("合计已发米", t["已发米"], 490.0)

print("② 成品库存表")
frows = models.list_finished(cid, "", False)
p2 = excel.export_finished(frows, customer="测试客户")
check("文件名带客户名", os.path.basename(p2), "成品库存_测试客户.xlsx")
ws2 = load_workbook(p2).active
hdr2, d2 = sheet_dict(ws2, 3, 4)
check("加工日期", d2["加工日期"], "2026-01-08")
check("缸号", d2["缸号"], "D001")
check("工艺", d2["工艺"], proc_name)
check("成品卷", d2["成品卷"], 3)
check("成品米", d2["成品米"], 500.0)
check("已发卷", d2["已发卷"], 3)
check("已发米", d2["已发米"], 490.0)
check("待发米（缩率零头）", d2["待发米"], 10.0)
check("状态", d2["状态"], "已发完")

print("③ 对账单")
st = services.statement(cid, None, None)
p3 = excel.export_statement(st)
check("文件生成", os.path.exists(p3), True)
ws3 = load_workbook(p3).active
hdr3, d3 = sheet_dict(ws3, 3, 4)
check("明细金额 = 490 × 3.5", d3["金额"], 1715.0)
check("明细缸号", d3["缸号"], "D001")

print()
print("通过 %d，失败 %d" % (ok, fail))
print("导出目录：%s" % TMP)
sys.exit(1 if fail else 0)
